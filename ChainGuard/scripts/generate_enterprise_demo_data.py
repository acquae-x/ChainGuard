"""Generate a deterministic enterprise demo data pack for ChainGuard.

The generated company, people, transactions, and incidents are synthetic.
They are intended for demos, integration tests, and data-mapping workshops.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


SEED = 20260615
AS_OF = datetime(2026, 6, 15, 8, 0, tzinfo=timezone(timedelta(hours=8)))
COMPANY = "华东智造科技有限公司"
# Historical outcome generation constants. They intentionally keep moderate
# noise so the demo data is learnable but not perfectly separable.
OUTCOME_SIGNAL_NOISE_STD = 0.72
OUTCOME_SUCCESS_THRESHOLD = -1.08
OUTCOME_PARTIAL_THRESHOLD = -1.86
QUALITY_COVERAGE_WEIGHT = 1.8
QUALITY_DELAY_WEIGHT = 1.2
QUALITY_COST_WEIGHT = 1.0
QUALITY_DOWNTIME_WEIGHT = 0.8
RESOURCE_TABLES = {
    "materials": "materials",
    "warehouses": "warehouses",
    "suppliers": "suppliers",
    "supplier-materials": "supplier_materials",
    "customers": "customers",
    "inventory": "inventory",
    "inventory-snapshots": "inventory_snapshots",
    "inventory-movements": "inventory_movements",
    "sales-orders": "sales_orders",
    "sales-order-lines": "sales_order_lines",
    "purchase-orders": "purchase_orders",
    "purchase-order-lines": "purchase_order_lines",
    "shipments": "shipments",
    "quality-inspections": "quality_inspections",
    "production-plans": "production_plans",
    "supplier-performance": "supplier_performance",
    "disruption-events": "disruption_events",
    "historical-decisions": "historical_decisions",
}


def iso(value: datetime) -> str:
    return value.isoformat(timespec="seconds")


def money(value: float) -> float:
    return round(value, 2)


def clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))


def derive_outcome_quality(
    *,
    covered_demand_rate: float,
    actual_delay_hours: float,
    predicted_delay_hours: float,
    actual_cost: float,
    predicted_cost: float,
    production_downtime_hours: float,
    rng: random.Random,
) -> float:
    delay_ratio = min(actual_delay_hours / max(predicted_delay_hours, 1.0), 3.0)
    cost_ratio = min(actual_cost / max(predicted_cost, 1.0), 3.0)
    return (
        QUALITY_COVERAGE_WEIGHT * (covered_demand_rate - 0.75)
        - QUALITY_DELAY_WEIGHT * delay_ratio / 3.0
        - QUALITY_COST_WEIGHT * cost_ratio / 3.0
        - QUALITY_DOWNTIME_WEIGHT * (production_downtime_hours / 48.0)
        + rng.gauss(0.0, OUTCOME_SIGNAL_NOISE_STD)
    )


def outcome_from_quality(quality_score: float) -> str:
    if quality_score > OUTCOME_SUCCESS_THRESHOLD:
        return "success"
    if quality_score > OUTCOME_PARTIAL_THRESHOLD:
        return "partial_success"
    return "failed"


def rating_from_outcome(
    outcome_status: str,
    quality_score: float,
    rng: random.Random,
) -> int:
    if outcome_status == "success":
        raw = 4.05 + 0.20 * (quality_score - OUTCOME_SUCCESS_THRESHOLD) + rng.gauss(0.0, 0.85)
    elif outcome_status == "partial_success":
        raw = 3.25 + 0.22 * (quality_score - OUTCOME_PARTIAL_THRESHOLD) + rng.gauss(0.0, 0.90)
    else:
        raw = 2.35 + 0.18 * (quality_score - OUTCOME_PARTIAL_THRESHOLD) + rng.gauss(0.0, 0.90)
    return int(round(clamp(raw, 1.0, 5.0)))


def weighted_choice(rng: random.Random, options: list[tuple[Any, int]]) -> Any:
    return rng.choices(
        [item for item, _ in options],
        weights=[weight for _, weight in options],
        k=1,
    )[0]


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"Cannot write empty dataset: {path}")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_data() -> dict[str, list[dict[str, Any]]]:
    rng = random.Random(SEED)
    regions = ["上海", "苏州", "宁波", "杭州", "合肥", "武汉", "东莞", "成都", "青岛", "天津"]
    categories = [
        ("控制芯片", "PCS", 0.18, (80, 680)),
        ("功率模块", "PCS", 0.15, (120, 1200)),
        ("传感器", "PCS", 0.15, (35, 360)),
        ("连接器", "PCS", 0.12, (2, 45)),
        ("结构件", "PCS", 0.12, (18, 220)),
        ("包装材料", "SET", 0.08, (1, 18)),
        ("化工辅料", "KG", 0.08, (12, 95)),
        ("电机组件", "PCS", 0.12, (90, 980)),
    ]
    warehouses = [
        {
            "warehouse_id": f"WH-{index:02d}",
            "warehouse_name": name,
            "region": region,
            "warehouse_type": warehouse_type,
            "capacity_units": capacity,
            "manager": manager,
            "status": "active",
            "updated_at": iso(AS_OF),
        }
        for index, (name, region, warehouse_type, capacity, manager) in enumerate(
            [
                ("上海中央仓", "上海", "central", 500000, "张敏"),
                ("苏州工厂仓", "苏州", "factory", 360000, "李强"),
                ("宁波保税仓", "宁波", "bonded", 260000, "王璐"),
                ("合肥工厂仓", "合肥", "factory", 320000, "赵晨"),
                ("武汉区域仓", "武汉", "regional", 220000, "陈杰"),
                ("东莞区域仓", "东莞", "regional", 240000, "刘婷"),
            ],
            start=1,
        )
    ]

    materials: list[dict[str, Any]] = []
    for index in range(1, 241):
        category, unit, _, cost_range = weighted_choice(
            rng, [(item, int(item[2] * 100)) for item in categories]
        )
        unit_cost = rng.uniform(*cost_range)
        criticality = weighted_choice(rng, [("A", 20), ("B", 45), ("C", 35)])
        daily_usage = rng.randint(20, 900) * (2 if criticality == "A" else 1)
        materials.append(
            {
                "material_id": f"MAT-{index:04d}",
                "material_name": f"{category}-{index:04d}",
                "category": category,
                "unit": unit,
                "criticality": criticality,
                "standard_cost": money(unit_cost),
                "sales_price": money(unit_cost * rng.uniform(1.25, 1.85)),
                "daily_consumption": daily_usage,
                "safety_stock_days": {"A": 14, "B": 9, "C": 5}[criticality],
                "shelf_life_days": rng.choice([365, 730, 1095, 1825]),
                "active": 1,
                "updated_at": iso(AS_OF - timedelta(minutes=rng.randint(0, 1440))),
            }
        )

    suppliers: list[dict[str, Any]] = []
    for index in range(1, 61):
        region = regions[(index - 1) % len(regions)]
        reliability = round(rng.uniform(72, 98), 1)
        suppliers.append(
            {
                "supplier_id": f"SUP-{index:03d}",
                "supplier_name": f"{region}协同供应商{index:03d}",
                "region": region,
                "supplier_tier": weighted_choice(rng, [("strategic", 20), ("preferred", 45), ("approved", 35)]),
                "reliability_score": reliability,
                "quality_score": round(rng.uniform(76, 99), 1),
                "financial_risk": weighted_choice(rng, [("low", 70), ("medium", 25), ("high", 5)]),
                "contact_name": f"供应商联系人{index:03d}",
                "contact_phone": f"138{index:08d}"[-11:],
                "status": "active" if reliability >= 76 else "watchlist",
                "updated_at": iso(AS_OF - timedelta(days=rng.randint(0, 30))),
            }
        )

    customers: list[dict[str, Any]] = []
    customer_industries = ["汽车零部件", "工业自动化", "新能源", "医疗设备", "消费电子", "轨道交通"]
    for index in range(1, 121):
        level = weighted_choice(rng, [("A", 18), ("B", 42), ("C", 40)])
        customers.append(
            {
                "customer_id": f"CUS-{index:04d}",
                "customer_name": f"{rng.choice(regions)}客户企业{index:04d}",
                "industry": rng.choice(customer_industries),
                "customer_level": level,
                "credit_limit": money(rng.uniform(200000, 8000000)),
                "payment_terms_days": rng.choice([0, 15, 30, 45, 60, 90]),
                "sla_hours": {"A": 48, "B": 72, "C": 120}[level],
                "penalty_rate": {"A": 0.035, "B": 0.018, "C": 0.008}[level],
                "region": rng.choice(regions),
                "status": "active",
                "updated_at": iso(AS_OF - timedelta(days=rng.randint(0, 30))),
            }
        )

    supplier_materials: list[dict[str, Any]] = []
    material_supplier_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for material in materials:
        selected = rng.sample(suppliers, rng.randint(3, 6))
        for rank, supplier in enumerate(selected, start=1):
            record = {
                "supplier_material_id": f"SM-{material['material_id'][4:]}-{rank:02d}",
                "supplier_id": supplier["supplier_id"],
                "material_id": material["material_id"],
                "supplier_rank": rank,
                "lead_time_hours": rng.choice([24, 36, 48, 72, 96, 120, 168]),
                "minimum_order_qty": rng.choice([50, 100, 200, 500, 1000]),
                "available_emergency_qty": rng.randint(300, 15000),
                "unit_cost": money(material["standard_cost"] * rng.uniform(0.92, 1.32)),
                "emergency_cost_multiplier": round(rng.uniform(1.05, 1.65), 2),
                "qualified": 1 if supplier["quality_score"] >= 80 else 0,
                "updated_at": iso(AS_OF - timedelta(hours=rng.randint(0, 240))),
            }
            supplier_materials.append(record)
            material_supplier_map[material["material_id"]].append(record)

    inventory: list[dict[str, Any]] = []
    inventory_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for material in materials:
        for warehouse in warehouses:
            daily = material["daily_consumption"]
            target = daily * material["safety_stock_days"] / len(warehouses)
            ratio = rng.uniform(0.2, 2.4)
            current = max(0, int(target * ratio))
            allocated = int(current * rng.uniform(0.0, 0.35))
            in_transit = int(target * rng.uniform(0.0, 1.2))
            record = {
                "inventory_id": f"INV-{material['material_id'][4:]}-{warehouse['warehouse_id'][3:]}",
                "material_id": material["material_id"],
                "warehouse_id": warehouse["warehouse_id"],
                "on_hand_qty": current,
                "allocated_qty": allocated,
                "available_qty": max(0, current - allocated),
                "in_transit_qty": in_transit,
                "quality_hold_qty": int(current * rng.uniform(0, 0.04)),
                "safety_stock_qty": int(target),
                "reorder_point_qty": int(target * 1.35),
                "average_unit_cost": material["standard_cost"],
                "last_counted_at": iso(AS_OF - timedelta(days=rng.randint(0, 30))),
                "updated_at": iso(AS_OF - timedelta(minutes=rng.randint(0, 180))),
            }
            inventory.append(record)
            inventory_lookup[(material["material_id"], warehouse["warehouse_id"])] = record

    sales_orders: list[dict[str, Any]] = []
    sales_order_lines: list[dict[str, Any]] = []
    start = AS_OF - timedelta(days=365)
    for index in range(1, 3501):
        customer = rng.choice(customers)
        created = start + timedelta(minutes=rng.randint(0, 365 * 24 * 60))
        status = weighted_choice(
            rng,
            [("delivered", 42), ("shipped", 16), ("confirmed", 20), ("at_risk", 12), ("cancelled", 5), ("overdue", 5)],
        )
        line_count = rng.randint(1, 5)
        total = 0.0
        total_cost = 0.0
        order_id = f"SO-{created:%Y%m}-{index:06d}"
        for line_no, material in enumerate(rng.sample(materials, line_count), start=1):
            quantity = rng.randint(10, 4500)
            unit_price = material["sales_price"] * rng.uniform(0.92, 1.12)
            line_total = quantity * unit_price
            total += line_total
            total_cost += quantity * material["standard_cost"]
            sales_order_lines.append(
                {
                    "sales_order_line_id": f"{order_id}-{line_no:02d}",
                    "sales_order_id": order_id,
                    "line_no": line_no,
                    "material_id": material["material_id"],
                    "warehouse_id": rng.choice(warehouses)["warehouse_id"],
                    "ordered_qty": quantity,
                    "allocated_qty": quantity if status in {"shipped", "delivered"} else int(quantity * rng.uniform(0.2, 1.0)),
                    "shipped_qty": quantity if status == "delivered" else int(quantity * rng.uniform(0, 0.9)),
                    "unit_price": money(unit_price),
                    "line_amount": money(line_total),
                    "requested_delivery_at": iso(created + timedelta(hours=customer["sla_hours"])),
                    "updated_at": iso(min(AS_OF, created + timedelta(days=rng.randint(0, 30)))),
                }
            )
        due_at = created + timedelta(hours=customer["sla_hours"] + rng.choice([-12, 0, 12, 24, 48]))
        sales_orders.append(
            {
                "sales_order_id": order_id,
                "customer_id": customer["customer_id"],
                "customer_level": customer["customer_level"],
                "order_created_at": iso(created),
                "promised_delivery_at": iso(due_at),
                "order_status": status,
                "currency": "CNY",
                "order_amount": money(total),
                "gross_profit": money(total - total_cost),
                "penalty_cost": money(total * customer["penalty_rate"]),
                "sales_region": customer["region"],
                "line_count": line_count,
                "updated_at": iso(min(AS_OF, created + timedelta(days=rng.randint(0, 45)))),
            }
        )

    purchase_orders: list[dict[str, Any]] = []
    purchase_order_lines: list[dict[str, Any]] = []
    for index in range(1, 1801):
        supplier = rng.choice(suppliers)
        created = start + timedelta(minutes=rng.randint(0, 365 * 24 * 60))
        status = weighted_choice(
            rng,
            [("received", 45), ("in_transit", 18), ("confirmed", 17), ("delayed", 12), ("draft", 5), ("cancelled", 3)],
        )
        candidates = [item for item in supplier_materials if item["supplier_id"] == supplier["supplier_id"]]
        selected = rng.sample(candidates, min(len(candidates), rng.randint(1, 4)))
        po_id = f"PO-{created:%Y%m}-{index:06d}"
        total = 0.0
        max_lead = 0
        for line_no, mapping in enumerate(selected, start=1):
            quantity = rng.randint(100, 12000)
            unit_cost = mapping["unit_cost"]
            line_total = quantity * unit_cost
            total += line_total
            max_lead = max(max_lead, mapping["lead_time_hours"])
            purchase_order_lines.append(
                {
                    "purchase_order_line_id": f"{po_id}-{line_no:02d}",
                    "purchase_order_id": po_id,
                    "line_no": line_no,
                    "material_id": mapping["material_id"],
                    "ordered_qty": quantity,
                    "received_qty": quantity if status == "received" else int(quantity * rng.uniform(0, 0.8)),
                    "unit_cost": unit_cost,
                    "line_amount": money(line_total),
                    "warehouse_id": rng.choice(warehouses)["warehouse_id"],
                    "updated_at": iso(min(AS_OF, created + timedelta(days=rng.randint(0, 45)))),
                }
            )
        expected = created + timedelta(hours=max_lead or 72)
        purchase_orders.append(
            {
                "purchase_order_id": po_id,
                "supplier_id": supplier["supplier_id"],
                "order_created_at": iso(created),
                "expected_arrival_at": iso(expected),
                "purchase_status": status,
                "currency": "CNY",
                "order_amount": money(total),
                "payment_terms_days": rng.choice([15, 30, 45, 60, 90]),
                "buyer": rng.choice(["采购一组", "采购二组", "战略采购组"]),
                "line_count": len(selected),
                "updated_at": iso(min(AS_OF, created + timedelta(days=rng.randint(0, 45)))),
            }
        )

    shipments: list[dict[str, Any]] = []
    shippable_pos = [item for item in purchase_orders if item["purchase_status"] != "draft"]
    for index in range(1, 3001):
        po = rng.choice(shippable_pos)
        planned = datetime.fromisoformat(po["expected_arrival_at"])
        delay = weighted_choice(rng, [(0, 54), (6, 12), (12, 11), (24, 10), (48, 8), (72, 5)])
        status = weighted_choice(
            rng, [("delivered", 45), ("in_transit", 29), ("delayed", 16), ("customs_hold", 6), ("blocked", 4)]
        )
        mode = weighted_choice(rng, [("road", 48), ("air", 12), ("rail", 18), ("sea", 22)])
        shipments.append(
            {
                "shipment_id": f"SHP-{index:07d}",
                "purchase_order_id": po["purchase_order_id"],
                "supplier_id": po["supplier_id"],
                "transport_mode": mode,
                "carrier": f"承运商{rng.randint(1, 24):02d}",
                "origin": rng.choice(regions),
                "destination_warehouse_id": rng.choice(warehouses)["warehouse_id"],
                "planned_arrival_at": iso(planned),
                "estimated_arrival_at": iso(planned + timedelta(hours=delay)),
                "delay_hours": delay,
                "shipment_status": status,
                "tracking_number": f"CG{index:012d}",
                "freight_cost": money(rng.uniform(1200, 68000) * (2.4 if mode == "air" else 1)),
                "updated_at": iso(AS_OF - timedelta(minutes=rng.randint(0, 1440))),
            }
        )

    inventory_movements: list[dict[str, Any]] = []
    movement_types = [("receipt", 28), ("issue", 36), ("transfer_in", 12), ("transfer_out", 12), ("adjustment", 7), ("quality_hold", 5)]
    for index in range(1, 35001):
        material = rng.choice(materials)
        warehouse = rng.choice(warehouses)
        movement_type = weighted_choice(rng, movement_types)
        quantity = rng.randint(1, 8000)
        if movement_type in {"issue", "transfer_out", "quality_hold"}:
            quantity *= -1
        movement_at = start + timedelta(minutes=rng.randint(0, 365 * 24 * 60))
        inventory_movements.append(
            {
                "movement_id": f"MOV-{index:08d}",
                "material_id": material["material_id"],
                "warehouse_id": warehouse["warehouse_id"],
                "movement_type": movement_type,
                "quantity": quantity,
                "unit_cost": material["standard_cost"],
                "reference_type": rng.choice(["sales_order", "purchase_order", "production_order", "manual"]),
                "reference_id": f"REF-{rng.randint(1, 999999):06d}",
                "movement_at": iso(movement_at),
                "operator": f"仓储用户{rng.randint(1, 45):02d}",
            }
        )

    inventory_snapshots: list[dict[str, Any]] = []
    for day_offset in range(31):
        snapshot_at = AS_OF - timedelta(days=day_offset)
        for item in inventory:
            factor = rng.uniform(0.78, 1.22)
            on_hand = max(0, int(item["on_hand_qty"] * factor))
            allocated = min(on_hand, int(item["allocated_qty"] * rng.uniform(0.75, 1.2)))
            inventory_snapshots.append(
                {
                    "snapshot_id": f"SNP-{snapshot_at:%Y%m%d}-{item['inventory_id'][4:]}",
                    "snapshot_date": snapshot_at.date().isoformat(),
                    "material_id": item["material_id"],
                    "warehouse_id": item["warehouse_id"],
                    "on_hand_qty": on_hand,
                    "available_qty": on_hand - allocated,
                    "allocated_qty": allocated,
                    "in_transit_qty": max(0, int(item["in_transit_qty"] * rng.uniform(0.7, 1.3))),
                    "safety_stock_qty": item["safety_stock_qty"],
                    "inventory_value": money(on_hand * item["average_unit_cost"]),
                }
            )

    disruption_events: list[dict[str, Any]] = []
    event_types = [
        ("typhoon", "台风天气"),
        ("port_shutdown", "港口停运"),
        ("supplier_shutdown", "供应商停产"),
        ("quality_recall", "质量召回"),
        ("route_blockage", "运输节点中断"),
        ("demand_surge", "需求突增"),
        ("power_shortage", "区域限电"),
        ("customs_delay", "海关查验延误"),
    ]
    for index in range(1, 241):
        event_type, label = rng.choice(event_types)
        started = start + timedelta(hours=rng.randint(0, 365 * 24))
        severity = weighted_choice(rng, [("low", 18), ("medium", 42), ("high", 30), ("critical", 10)])
        risk_score = {
            "low": rng.randint(25, 49),
            "medium": rng.randint(50, 69),
            "high": rng.randint(70, 84),
            "critical": rng.randint(85, 98),
        }[severity]
        supplier = rng.choice(suppliers)
        material = rng.choice(materials)
        disruption_events.append(
            {
                "event_id": f"EVT-{index:06d}",
                "event_type": event_type,
                "event_title": f"{label}-{rng.choice(regions)}-{index:03d}",
                "severity": severity,
                "risk_score": risk_score,
                "location": rng.choice(regions),
                "affected_supplier_id": supplier["supplier_id"],
                "affected_material_id": material["material_id"],
                "affected_route": f"{supplier['region']}至{rng.choice(regions)}",
                "estimated_delay_hours": rng.choice([6, 12, 24, 36, 48, 72, 96, 168]),
                "event_status": weighted_choice(rng, [("resolved", 52), ("monitoring", 23), ("active", 20), ("escalated", 5)]),
                "started_at": iso(started),
                "resolved_at": iso(started + timedelta(hours=rng.randint(6, 240))) if started < AS_OF - timedelta(days=10) else "",
                "description": f"{label}影响供应商、物料和运输路线，需评估库存、订单与替代供应。",
                "updated_at": iso(min(AS_OF, started + timedelta(hours=rng.randint(1, 72)))),
            }
        )

    historical_decisions: list[dict[str, Any]] = []
    strategies = [
        "关键订单空运+备用供应商补货",
        "库存重分配+客户分级沟通",
        "铁路替代+延期交付",
        "双供应商分单+安全库存锁定",
        "紧急采购+生产计划调整",
        "接受短期停线+控制现金成本",
    ]
    for index in range(1, 601):
        event = rng.choice(disruption_events)
        strategy = rng.choice(strategies)
        latent_quality = rng.gauss(0.0, 1.0)
        predicted_delay = rng.choice([6, 12, 24, 36, 48, 72])
        predicted_cost = money(rng.uniform(20000, 900000))
        covered_demand_rate = round(
            clamp(0.77 + 0.13 * latent_quality + rng.gauss(0.0, 0.08), 0.52, 1.0),
            4,
        )
        delay_ratio = clamp(1.05 - 0.34 * latent_quality + rng.gauss(0.0, 0.22), 0.18, 2.80)
        cost_ratio = clamp(1.06 - 0.26 * latent_quality + rng.gauss(0.0, 0.20), 0.42, 2.60)
        production_downtime = int(
            round(clamp(18.0 - 8.5 * latent_quality + rng.gauss(0.0, 8.0), 0.0, 72.0))
        )
        actual_delay = round(max(0.0, predicted_delay * delay_ratio), 2)
        actual_cost = money(predicted_cost * cost_ratio)
        quality_score = derive_outcome_quality(
            covered_demand_rate=covered_demand_rate,
            actual_delay_hours=actual_delay,
            predicted_delay_hours=predicted_delay,
            actual_cost=actual_cost,
            predicted_cost=predicted_cost,
            production_downtime_hours=production_downtime,
            rng=rng,
        )
        outcome_status = outcome_from_quality(quality_score)
        human_rating = rating_from_outcome(outcome_status, quality_score, rng)
        failure_pressure = clamp(
            (OUTCOME_SUCCESS_THRESHOLD - quality_score + 0.85)
            + (1.0 - covered_demand_rate) * 2.0
            + production_downtime / 72.0,
            0.0,
            3.0,
        )
        lost_orders = int(round(clamp(rng.gauss(failure_pressure * 1.15, 0.85), 0.0, 8.0)))
        customer_complaints = int(
            round(clamp(rng.gauss(failure_pressure * 4.5 + lost_orders, 2.2), 0.0, 24.0))
        )
        historical_decisions.append(
            {
                "case_id": f"CASE-{index:06d}",
                "event_id": event["event_id"],
                "scenario": event["event_title"],
                "selected_strategy": strategy,
                "predicted_delay_hours": predicted_delay,
                "actual_delay_hours": actual_delay,
                "predicted_cost": predicted_cost,
                "actual_cost": actual_cost,
                "covered_demand_rate": covered_demand_rate,
                "production_downtime_hours": production_downtime,
                "lost_orders": lost_orders,
                "customer_complaints": customer_complaints,
                "outcome_status": outcome_status,
                "human_rating": human_rating,
                "lessons_learned": "优先保障A级客户，并同时验证备用供应商产能、质量与运输可用性。",
                "model_version": "demo-rule-v1",
                "parameter_version": "expert-2026-06",
                "created_at": iso(AS_OF - timedelta(days=rng.randint(1, 365))),
            }
        )

    quality_inspections: list[dict[str, Any]] = []
    for index in range(1, 2001):
        supplier = rng.choice(suppliers)
        mapping = rng.choice([item for item in supplier_materials if item["supplier_id"] == supplier["supplier_id"]])
        inspected = rng.randint(50, 5000)
        defect_rate = max(0.0, rng.gauss((100 - supplier["quality_score"]) / 100, 0.015))
        defects = min(inspected, int(inspected * defect_rate))
        quality_inspections.append(
            {
                "inspection_id": f"QI-{index:07d}",
                "supplier_id": supplier["supplier_id"],
                "material_id": mapping["material_id"],
                "purchase_order_id": rng.choice(purchase_orders)["purchase_order_id"],
                "inspected_qty": inspected,
                "defect_qty": defects,
                "defect_rate": round(defects / inspected, 5),
                "result": "pass" if defects / inspected <= 0.03 else "fail",
                "inspection_type": rng.choice(["incoming", "process", "final"]),
                "inspected_at": iso(start + timedelta(hours=rng.randint(0, 365 * 24))),
                "inspector": f"质检员{rng.randint(1, 20):02d}",
            }
        )

    production_plans: list[dict[str, Any]] = []
    for index in range(1, 2001):
        material = rng.choice(materials)
        planned_start = start + timedelta(hours=rng.randint(0, 400 * 24))
        status = weighted_choice(rng, [("completed", 42), ("released", 20), ("planned", 22), ("at_risk", 11), ("paused", 5)])
        production_plans.append(
            {
                "production_plan_id": f"PP-{index:07d}",
                "plant_warehouse_id": rng.choice([warehouses[1], warehouses[3]])["warehouse_id"],
                "finished_material_id": material["material_id"],
                "planned_qty": rng.randint(100, 20000),
                "planned_start_at": iso(planned_start),
                "planned_end_at": iso(planned_start + timedelta(hours=rng.randint(8, 120))),
                "plan_status": status,
                "priority": weighted_choice(rng, [("urgent", 12), ("high", 28), ("normal", 60)]),
                "material_readiness_rate": round(rng.uniform(0.55, 1.0), 4),
                "updated_at": iso(min(AS_OF, planned_start + timedelta(hours=rng.randint(0, 48)))),
            }
        )

    supplier_performance: list[dict[str, Any]] = []
    for supplier in suppliers:
        for month_offset in range(12):
            month = (AS_OF.replace(day=1) - timedelta(days=month_offset * 28)).strftime("%Y-%m")
            on_time = min(1.0, max(0.55, supplier["reliability_score"] / 100 + rng.uniform(-0.08, 0.04)))
            defect = min(0.12, max(0.002, (100 - supplier["quality_score"]) / 100 + rng.uniform(-0.01, 0.015)))
            supplier_performance.append(
                {
                    "supplier_performance_id": f"SP-{supplier['supplier_id'][4:]}-{month}",
                    "supplier_id": supplier["supplier_id"],
                    "period": month,
                    "purchase_order_count": rng.randint(4, 65),
                    "on_time_delivery_rate": round(on_time, 4),
                    "defect_rate": round(defect, 5),
                    "average_delay_hours": round(max(0, rng.gauss((1 - on_time) * 80, 6)), 1),
                    "emergency_response_hours": round(rng.uniform(1, 36), 1),
                    "score": round((on_time * 65 + (1 - defect) * 35), 1),
                }
            )

    return {
        "materials": materials,
        "warehouses": warehouses,
        "suppliers": suppliers,
        "supplier_materials": supplier_materials,
        "customers": customers,
        "inventory": inventory,
        "inventory_snapshots": inventory_snapshots,
        "inventory_movements": inventory_movements,
        "sales_orders": sales_orders,
        "sales_order_lines": sales_order_lines,
        "purchase_orders": purchase_orders,
        "purchase_order_lines": purchase_order_lines,
        "shipments": shipments,
        "quality_inspections": quality_inspections,
        "production_plans": production_plans,
        "supplier_performance": supplier_performance,
        "disruption_events": disruption_events,
        "historical_decisions": historical_decisions,
    }


def sqlite_type(values: Iterable[Any]) -> str:
    for value in values:
        if value is None or value == "":
            continue
        if isinstance(value, bool) or isinstance(value, int):
            return "INTEGER"
        if isinstance(value, float):
            return "REAL"
        return "TEXT"
    return "TEXT"


def write_sqlite(path: Path, datasets: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path, isolation_level=None)
    try:
        # Some demo sandboxes block SQLite sidecar journal files.
        connection.execute("PRAGMA journal_mode=OFF")
        connection.execute("PRAGMA synchronous=OFF")
        connection.execute("PRAGMA temp_store=MEMORY")
        for table_name, rows in datasets.items():
            columns = list(rows[0])
            definitions = ", ".join(
                f'"{column}" {sqlite_type(row[column] for row in rows[:100])}' for column in columns
            )
            connection.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            connection.execute(f'CREATE TABLE "{table_name}" ({definitions})')
            placeholders = ", ".join("?" for _ in columns)
            connection.executemany(
                f'INSERT INTO "{table_name}" VALUES ({placeholders})',
                [[row[column] for column in columns] for row in rows],
            )
        connection.execute("CREATE INDEX idx_inventory_material ON inventory(material_id)")
        connection.execute("CREATE INDEX idx_so_customer ON sales_orders(customer_id)")
        connection.execute("CREATE INDEX idx_shipment_status ON shipments(shipment_status)")
        connection.execute("CREATE INDEX idx_event_status ON disruption_events(event_status)")
    finally:
        connection.close()


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ws.columns:
        values = [str(cell.value or "") for cell in list(column)[:200]]
        width = min(34, max(10, max(len(value) for value in values) + 2))
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def add_rows(ws, rows: list[dict[str, Any]]) -> None:
    headers = list(rows[0])
    ws.append(headers)
    for row in rows:
        ws.append([row[header] for header in headers])
    style_sheet(ws)


def write_small_business_xlsx(path: Path, data: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "经营看板"
    dashboard.sheet_view.showGridLines = False
    dashboard["A1"] = "ChainGuard 小企业供应链演示账套"
    dashboard["A1"].font = Font(size=20, bold=True, color="17365D")
    dashboard.merge_cells("A1:F1")
    dashboard["A3"] = "统计时点"
    dashboard["B3"] = iso(AS_OF)
    dashboard["A4"] = "企业"
    dashboard["B4"] = COMPANY

    inventory_rows = sorted(
        data["inventory"],
        key=lambda row: row["available_qty"] - row["safety_stock_qty"],
    )[:180]
    sales_rows = data["sales_orders"][-800:]
    purchase_rows = data["purchase_orders"][-350:]
    supplier_rows = data["suppliers"][:35]
    customer_rows = data["customers"][:80]
    event_rows = data["disruption_events"][-80:]
    cashflow_rows = []
    for index in range(1, 501):
        source = sales_rows[index % len(sales_rows)] if index % 2 else purchase_rows[index % len(purchase_rows)]
        is_income = index % 2 == 1
        cashflow_rows.append(
            {
                "流水号": f"CF-{index:06d}",
                "日期": (AS_OF.date() - timedelta(days=index % 180)).isoformat(),
                "类型": "销售回款" if is_income else "采购付款",
                "业务单号": source["sales_order_id"] if is_income else source["purchase_order_id"],
                "收入": source["order_amount"] if is_income else 0,
                "支出": 0 if is_income else source["order_amount"],
                "账户": "基本户",
                "状态": "已入账" if index % 7 else "待核销",
            }
        )

    sheets = {
        "库存台账": inventory_rows,
        "销售订单": sales_rows,
        "采购订单": purchase_rows,
        "供应商": supplier_rows,
        "客户": customer_rows,
        "资金流水": cashflow_rows,
        "风险事件": event_rows,
    }
    for sheet_name, rows in sheets.items():
        add_rows(workbook.create_sheet(sheet_name), rows)

    dictionary = workbook.create_sheet("字段说明")
    dictionary_rows = [
        {"工作表": "库存台账", "用途": "当前库存、锁定量、安全库存和在途量"},
        {"工作表": "销售订单", "用途": "客户订单、交期、毛利和延期罚金"},
        {"工作表": "采购订单", "用途": "供应商采购、到货计划和订单状态"},
        {"工作表": "供应商", "用途": "可靠性、质量和财务风险"},
        {"工作表": "客户", "用途": "客户分级、信用额度和服务等级"},
        {"工作表": "资金流水", "用途": "小企业现金流演示数据"},
        {"工作表": "风险事件", "用途": "供应链异常事件与风险评分"},
    ]
    add_rows(dictionary, dictionary_rows)

    metrics = [
        ("库存记录", len(inventory_rows)),
        ("库存总价值", sum(row["on_hand_qty"] * row["average_unit_cost"] for row in inventory_rows)),
        ("低于安全库存", sum(row["available_qty"] < row["safety_stock_qty"] for row in inventory_rows)),
        ("销售订单", len(sales_rows)),
        ("风险订单", sum(row["order_status"] in {"at_risk", "overdue"} for row in sales_rows)),
        ("采购订单", len(purchase_rows)),
        ("延误采购", sum(row["purchase_status"] == "delayed" for row in purchase_rows)),
        ("活动风险事件", sum(row["event_status"] in {"active", "escalated"} for row in event_rows)),
    ]
    dashboard.append([])
    dashboard.append(["指标", "数值"])
    for label, value in metrics:
        dashboard.append([label, value])
    for cell in dashboard[6]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    dashboard.column_dimensions["A"].width = 24
    dashboard.column_dimensions["B"].width = 22
    dashboard["B8"].number_format = '¥#,##0.00'

    event_counter = Counter(row["severity"] for row in event_rows)
    dashboard["D6"] = "风险等级"
    dashboard["E6"] = "数量"
    for row_index, severity in enumerate(["low", "medium", "high", "critical"], start=7):
        dashboard.cell(row_index, 4, severity)
        dashboard.cell(row_index, 5, event_counter[severity])
    pie = PieChart()
    pie.title = "风险事件等级分布"
    pie.add_data(Reference(dashboard, min_col=5, min_row=6, max_row=10), titles_from_data=True)
    pie.set_categories(Reference(dashboard, min_col=4, min_row=7, max_row=10))
    dashboard.add_chart(pie, "D12")

    bar = BarChart()
    bar.title = "核心经营指标"
    bar.add_data(Reference(dashboard, min_col=2, min_row=6, max_row=13), titles_from_data=True)
    bar.set_categories(Reference(dashboard, min_col=1, min_row=7, max_row=13))
    dashboard.add_chart(bar, "G3")
    workbook.save(path)


def register_chinese_font() -> str:
    candidates = [
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
    ]
    for candidate in candidates:
        if candidate.exists():
            pdfmetrics.registerFont(TTFont("ChainGuardCN", str(candidate)))
            return "ChainGuardCN"
    return "Helvetica"


def pdf_styles(font_name: str):
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            "CGTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=24,
            leading=32,
            textColor=colors.HexColor("#17365D"),
            alignment=TA_CENTER,
            spaceAfter=12,
        )
    )
    styles.add(
        ParagraphStyle(
            "CGHeading",
            parent=styles["Heading2"],
            fontName=font_name,
            fontSize=15,
            leading=20,
            textColor=colors.HexColor("#17365D"),
            spaceBefore=8,
            spaceAfter=8,
        )
    )
    styles.add(
        ParagraphStyle(
            "CGBody",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=9,
            leading=14,
        )
    )
    styles.add(
        ParagraphStyle(
            "CGSmall",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=7,
            leading=10,
        )
    )
    styles.add(
        ParagraphStyle(
            "CGRight",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=8,
            leading=11,
            alignment=TA_RIGHT,
        )
    )
    return styles


def page_footer(canvas, document) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#667085"))
    canvas.drawString(18 * mm, 10 * mm, "Synthetic demo data - not a real enterprise record")
    canvas.drawRightString(A4[0] - 18 * mm, 10 * mm, f"Page {document.page}")
    canvas.restoreState()


def make_table(rows: list[list[Any]], font_name: str, widths=None, repeat_rows: int = 1) -> Table:
    table = Table(rows, colWidths=widths, repeatRows=repeat_rows)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D5DD")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def write_management_report(path: Path, data: dict[str, list[dict[str, Any]]], font_name: str) -> None:
    styles = pdf_styles(font_name)
    document = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        title="ChainGuard 企业供应链风险管理周报",
        author=COMPANY,
    )
    risky_inventory = sorted(
        data["inventory"],
        key=lambda row: row["available_qty"] - row["safety_stock_qty"],
    )
    active_events = [
        row for row in data["disruption_events"] if row["event_status"] in {"active", "escalated"}
    ]
    at_risk_orders = [
        row for row in data["sales_orders"] if row["order_status"] in {"at_risk", "overdue"}
    ]
    delayed_shipments = [row for row in data["shipments"] if row["delay_hours"] > 0]
    latest_performance = [row for row in data["supplier_performance"] if row["period"] == "2026-06"]
    lowest_suppliers = sorted(latest_performance, key=lambda row: row["score"])[:15]

    story = [
        Spacer(1, 28 * mm),
        Paragraph("ChainGuard 企业供应链风险管理周报", styles["CGTitle"]),
        Paragraph(COMPANY, styles["CGTitle"]),
        Spacer(1, 8 * mm),
        Paragraph(f"统计时点：{AS_OF:%Y-%m-%d %H:%M}（UTC+8）", styles["CGBody"]),
        Paragraph("用途：比赛演示、ERP 接口联调、数据映射和算法回放", styles["CGBody"]),
        Spacer(1, 16 * mm),
        Paragraph(
            "声明：本报告全部企业、人员、订单、金额和事件均为程序生成的合成数据，不对应任何真实主体。",
            styles["CGBody"],
        ),
        PageBreak(),
        Paragraph("一、管理层摘要", styles["CGHeading"]),
        make_table(
            [
                ["指标", "数值", "管理含义"],
                ["物料数", f"{len(data['materials']):,}", "覆盖多品类制造企业主数据"],
                ["当前库存记录", f"{len(data['inventory']):,}", "按物料与仓库拆分"],
                ["低于安全库存", f"{sum(row['available_qty'] < row['safety_stock_qty'] for row in data['inventory']):,}", "建议进入风险排序"],
                ["风险/逾期销售订单", f"{len(at_risk_orders):,}", "需结合客户级别与罚金处理"],
                ["延误运输单", f"{len(delayed_shipments):,}", "需评估替代运输和在途改配"],
                ["活动/升级事件", f"{len(active_events):,}", "需人工确认应急决策"],
                ["历史决策案例", f"{len(data['historical_decisions']):,}", "用于规则回放和参数校准"],
            ],
            font_name,
            widths=[42 * mm, 30 * mm, 92 * mm],
        ),
        Spacer(1, 6 * mm),
        Paragraph(
            "本周建议优先处理A级物料安全库存缺口、A级客户风险订单和高风险供应商在途延误。"
            "演示时可从 ERP API 拉取同口径数据，再由 ChainGuard 进行风险计算、策略生成和历史案例检索。",
            styles["CGBody"],
        ),
        Paragraph("二、库存风险 Top 25", styles["CGHeading"]),
        make_table(
            [["物料", "仓库", "可用", "安全库存", "缺口", "在途"]]
            + [
                [
                    row["material_id"],
                    row["warehouse_id"],
                    f"{row['available_qty']:,}",
                    f"{row['safety_stock_qty']:,}",
                    f"{row['available_qty'] - row['safety_stock_qty']:,}",
                    f"{row['in_transit_qty']:,}",
                ]
                for row in risky_inventory[:25]
            ],
            font_name,
            widths=[28 * mm, 24 * mm, 27 * mm, 27 * mm, 27 * mm, 27 * mm],
        ),
        PageBreak(),
        Paragraph("三、风险订单 Top 20", styles["CGHeading"]),
        make_table(
            [["订单", "客户级别", "状态", "订单金额", "毛利", "延期罚金"]]
            + [
                [
                    row["sales_order_id"],
                    row["customer_level"],
                    row["order_status"],
                    f"¥{row['order_amount']:,.0f}",
                    f"¥{row['gross_profit']:,.0f}",
                    f"¥{row['penalty_cost']:,.0f}",
                ]
                for row in sorted(at_risk_orders, key=lambda item: item["penalty_cost"], reverse=True)[:20]
            ],
            font_name,
            widths=[37 * mm, 22 * mm, 23 * mm, 31 * mm, 27 * mm, 30 * mm],
        ),
        Spacer(1, 6 * mm),
        Paragraph("四、供应商绩效低分清单", styles["CGHeading"]),
        make_table(
            [["供应商", "期间", "准时率", "缺陷率", "平均延误", "得分"]]
            + [
                [
                    row["supplier_id"],
                    row["period"],
                    f"{row['on_time_delivery_rate']:.1%}",
                    f"{row['defect_rate']:.2%}",
                    f"{row['average_delay_hours']:.1f}h",
                    row["score"],
                ]
                for row in lowest_suppliers
            ],
            font_name,
            widths=[31 * mm, 24 * mm, 28 * mm, 28 * mm, 31 * mm, 24 * mm],
        ),
        PageBreak(),
        Paragraph("五、活动风险事件", styles["CGHeading"]),
        make_table(
            [["事件", "类型", "等级", "地点", "风险分", "延误"]]
            + [
                [
                    row["event_id"],
                    row["event_type"],
                    row["severity"],
                    row["location"],
                    row["risk_score"],
                    f"{row['estimated_delay_hours']}h",
                ]
                for row in sorted(active_events, key=lambda item: item["risk_score"], reverse=True)[:30]
            ],
            font_name,
            widths=[29 * mm, 38 * mm, 23 * mm, 23 * mm, 23 * mm, 23 * mm],
        ),
        Spacer(1, 6 * mm),
        Paragraph("六、数据覆盖与演示建议", styles["CGHeading"]),
        make_table(
            [["数据集", "行数", "典型来源"]]
            + [
                [name, f"{len(rows):,}", "ERP/WMS/TMS/SRM/质量/应急台账"]
                for name, rows in data.items()
            ],
            font_name,
            widths=[58 * mm, 35 * mm, 75 * mm],
        ),
    ]
    document.build(story, onFirstPage=page_footer, onLaterPages=page_footer)


def write_inventory_report(path: Path, data: dict[str, list[dict[str, Any]]], font_name: str) -> None:
    styles = pdf_styles(font_name)
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=16 * mm,
        title="库存与安全库存缺口明细",
    )
    material_lookup = {row["material_id"]: row for row in data["materials"]}
    rows = sorted(
        data["inventory"],
        key=lambda row: row["available_qty"] - row["safety_stock_qty"],
    )
    table_rows = [["物料", "名称", "等级", "仓库", "现有", "锁定", "可用", "安全库存", "缺口", "在途", "库存价值"]]
    for row in rows:
        material = material_lookup[row["material_id"]]
        table_rows.append(
            [
                row["material_id"],
                material["material_name"],
                material["criticality"],
                row["warehouse_id"],
                f"{row['on_hand_qty']:,}",
                f"{row['allocated_qty']:,}",
                f"{row['available_qty']:,}",
                f"{row['safety_stock_qty']:,}",
                f"{row['available_qty'] - row['safety_stock_qty']:,}",
                f"{row['in_transit_qty']:,}",
                f"¥{row['on_hand_qty'] * row['average_unit_cost']:,.0f}",
            ]
        )
    story = [
        Paragraph("库存与安全库存缺口明细", styles["CGTitle"]),
        Paragraph(f"{COMPANY} | {AS_OF:%Y-%m-%d %H:%M}", styles["CGBody"]),
        Spacer(1, 5 * mm),
        make_table(
            table_rows,
            font_name,
            widths=[22 * mm, 37 * mm, 13 * mm, 17 * mm, 19 * mm, 19 * mm, 19 * mm, 22 * mm, 20 * mm, 19 * mm, 29 * mm],
        ),
    ]
    document.build(story, onFirstPage=page_footer, onLaterPages=page_footer)


def write_purchase_order_pdf(
    path: Path,
    order: dict[str, Any],
    lines: list[dict[str, Any]],
    supplier: dict[str, Any],
    font_name: str,
) -> None:
    styles = pdf_styles(font_name)
    document = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=14 * mm, bottomMargin=18 * mm)
    story = [
        Paragraph("采购订单", styles["CGTitle"]),
        make_table(
            [
                ["采购单号", order["purchase_order_id"], "供应商", supplier["supplier_name"]],
                ["下单时间", order["order_created_at"], "预计到货", order["expected_arrival_at"]],
                ["采购组", order["buyer"], "状态", order["purchase_status"]],
            ],
            font_name,
            widths=[27 * mm, 55 * mm, 27 * mm, 60 * mm],
            repeat_rows=0,
        ),
        Spacer(1, 7 * mm),
        make_table(
            [["行", "物料", "数量", "已收", "单价", "金额", "仓库"]]
            + [
                [
                    line["line_no"],
                    line["material_id"],
                    f"{line['ordered_qty']:,}",
                    f"{line['received_qty']:,}",
                    f"¥{line['unit_cost']:,.2f}",
                    f"¥{line['line_amount']:,.2f}",
                    line["warehouse_id"],
                ]
                for line in lines
            ],
            font_name,
            widths=[14 * mm, 32 * mm, 24 * mm, 24 * mm, 27 * mm, 34 * mm, 20 * mm],
        ),
        Spacer(1, 8 * mm),
        Paragraph(f"订单总额：¥{order['order_amount']:,.2f}", styles["CGRight"]),
        Paragraph("交付要求：包装完整、批次可追溯；发生延期或质量异常时须在 4 小时内反馈。", styles["CGBody"]),
        Spacer(1, 20 * mm),
        make_table(
            [["采购审核", "财务审核", "供应商确认"], ["________________", "________________", "________________"]],
            font_name,
            widths=[58 * mm, 58 * mm, 58 * mm],
        ),
    ]
    document.build(story, onFirstPage=page_footer)


def write_delivery_note_pdf(
    path: Path,
    shipment: dict[str, Any],
    po_lines: list[dict[str, Any]],
    font_name: str,
) -> None:
    styles = pdf_styles(font_name)
    document = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=14 * mm, bottomMargin=18 * mm)
    story = [
        Paragraph("送货与在途跟踪单", styles["CGTitle"]),
        make_table(
            [
                ["运输单号", shipment["shipment_id"], "采购单号", shipment["purchase_order_id"]],
                ["承运商", shipment["carrier"], "运输方式", shipment["transport_mode"]],
                ["起点", shipment["origin"], "目的仓", shipment["destination_warehouse_id"]],
                ["计划到达", shipment["planned_arrival_at"], "预计到达", shipment["estimated_arrival_at"]],
                ["状态", shipment["shipment_status"], "延误小时", shipment["delay_hours"]],
            ],
            font_name,
            widths=[27 * mm, 58 * mm, 27 * mm, 58 * mm],
            repeat_rows=0,
        ),
        Spacer(1, 7 * mm),
        make_table(
            [["行", "物料", "发运数量", "目标仓库"]]
            + [
                [line["line_no"], line["material_id"], f"{line['ordered_qty']:,}", line["warehouse_id"]]
                for line in po_lines
            ],
            font_name,
            widths=[25 * mm, 50 * mm, 48 * mm, 48 * mm],
        ),
        Spacer(1, 12 * mm),
        Paragraph("异常处理：若预计延误超过 12 小时，需同步采购、物流和生产计划负责人。", styles["CGBody"]),
        Spacer(1, 18 * mm),
        make_table(
            [["发货确认", "承运商签字", "收货确认"], ["________________", "________________", "________________"]],
            font_name,
            widths=[58 * mm, 58 * mm, 58 * mm],
        ),
    ]
    document.build(story, onFirstPage=page_footer)


def write_event_brief_pdf(path: Path, event: dict[str, Any], font_name: str) -> None:
    styles = pdf_styles(font_name)
    document = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=20 * mm, leftMargin=20 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    story = [
        Paragraph("供应链风险事件简报", styles["CGTitle"]),
        make_table(
            [
                ["事件编号", event["event_id"], "风险分", event["risk_score"]],
                ["事件名称", event["event_title"], "等级", event["severity"]],
                ["事件类型", event["event_type"], "状态", event["event_status"]],
                ["地点", event["location"], "预计延误", f"{event['estimated_delay_hours']} 小时"],
                ["受影响供应商", event["affected_supplier_id"], "受影响物料", event["affected_material_id"]],
                ["受影响路线", event["affected_route"], "开始时间", event["started_at"]],
            ],
            font_name,
            widths=[30 * mm, 61 * mm, 30 * mm, 49 * mm],
            repeat_rows=0,
        ),
        Spacer(1, 8 * mm),
        Paragraph("事件描述", styles["CGHeading"]),
        Paragraph(event["description"], styles["CGBody"]),
        Paragraph("建议动作", styles["CGHeading"]),
        Paragraph("1. 核对受影响物料在各仓可用库存和关键订单需求。", styles["CGBody"]),
        Paragraph("2. 询价备用供应商，确认可供量、交期、质量资质和加急成本。", styles["CGBody"]),
        Paragraph("3. 比较空运、陆运、铁路等替代路线，并设置人工审批点。", styles["CGBody"]),
        Paragraph("4. 对高优先级客户建立主动沟通和交付承诺更新。", styles["CGBody"]),
        Paragraph("5. 事件关闭后记录实际成本、延期和客户影响，形成经验卡片。", styles["CGBody"]),
    ]
    document.build(story, onFirstPage=page_footer)


def write_pdfs(root: Path, data: dict[str, list[dict[str, Any]]]) -> None:
    root.mkdir(parents=True, exist_ok=True)
    font_name = register_chinese_font()
    write_management_report(root / "企业供应链风险管理周报.pdf", data, font_name)
    write_inventory_report(root / "库存与安全库存缺口明细.pdf", data, font_name)

    supplier_lookup = {row["supplier_id"]: row for row in data["suppliers"]}
    po_line_lookup: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for line in data["purchase_order_lines"]:
        po_line_lookup[line["purchase_order_id"]].append(line)

    po_dir = root / "业务单据" / "采购订单"
    delivery_dir = root / "业务单据" / "送货单"
    event_dir = root / "风险事件简报"
    po_dir.mkdir(parents=True, exist_ok=True)
    delivery_dir.mkdir(parents=True, exist_ok=True)
    event_dir.mkdir(parents=True, exist_ok=True)
    for order in data["purchase_orders"][-12:]:
        write_purchase_order_pdf(
            po_dir / f"{order['purchase_order_id']}.pdf",
            order,
            po_line_lookup[order["purchase_order_id"]],
            supplier_lookup[order["supplier_id"]],
            font_name,
        )
    for shipment in data["shipments"][-12:]:
        write_delivery_note_pdf(
            delivery_dir / f"{shipment['shipment_id']}.pdf",
            shipment,
            po_line_lookup[shipment["purchase_order_id"]],
            font_name,
        )
    active_events = sorted(data["disruption_events"], key=lambda row: row["risk_score"], reverse=True)[:8]
    for event in active_events:
        write_event_brief_pdf(event_dir / f"{event['event_id']}.pdf", event, font_name)


def write_api_files(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    resources = "\n".join(
        f"  /api/v1/{resource}:\n"
        "    get:\n"
        f"      summary: 分页查询 {resource}\n"
        "      parameters:\n"
        "        - in: query\n"
        "          name: page\n"
        "          schema: {type: integer, default: 1, minimum: 1}\n"
        "        - in: query\n"
        "          name: page_size\n"
        "          schema: {type: integer, default: 100, minimum: 1, maximum: 1000}\n"
        "      responses:\n"
        "        '200': {description: 成功}\n"
        for resource in RESOURCE_TABLES
    )
    openapi = (
        "openapi: 3.0.3\n"
        "info:\n"
        "  title: ChainGuard Synthetic ERP API\n"
        "  version: 1.0.0\n"
        "  description: 合成企业数据接口，仅用于演示和联调。\n"
        "servers:\n"
        "  - url: http://127.0.0.1:8765\n"
        "paths:\n"
        "  /health:\n"
        "    get:\n"
        "      summary: 健康检查\n"
        "      responses:\n"
        "        '200': {description: 服务正常}\n"
        "  /api/v1/catalog:\n"
        "    get:\n"
        "      summary: 获取资源目录与记录数\n"
        "      responses:\n"
        "        '200': {description: 成功}\n"
        "  /api/v1/dashboard/summary:\n"
        "    get:\n"
        "      summary: 获取供应链驾驶舱摘要\n"
        "      responses:\n"
        "        '200': {description: 成功}\n"
        f"{resources}"
    )
    (root / "openapi.yaml").write_text(openapi, encoding="utf-8")
    (root / "sample_requests.http").write_text(
        """### Health
GET http://127.0.0.1:8765/health

### Resource catalog
GET http://127.0.0.1:8765/api/v1/catalog

### Dashboard summary
GET http://127.0.0.1:8765/api/v1/dashboard/summary

### Inventory page
GET http://127.0.0.1:8765/api/v1/inventory?page=1&page_size=20

### Active disruption events
GET http://127.0.0.1:8765/api/v1/disruption-events?page=1&page_size=50&event_status=active

### At-risk orders
GET http://127.0.0.1:8765/api/v1/sales-orders?page=1&page_size=50&order_status=at_risk
""",
        encoding="utf-8",
    )


def build_manifest(root: Path, data: dict[str, list[dict[str, Any]]]) -> None:
    files = []
    for path in sorted(root.rglob("*")):
        if path.is_file():
            files.append(
                {
                    "path": path.relative_to(root).as_posix(),
                    "size_bytes": path.stat().st_size,
                }
            )
    manifest = {
        "dataset_name": "ChainGuard enterprise synthetic demo pack",
        "company": COMPANY,
        "generated_at": iso(AS_OF),
        "seed": SEED,
        "synthetic_data": True,
        "record_counts": {name: len(rows) for name, rows in data.items()},
        "total_records": sum(len(rows) for rows in data.values()),
        "files": files,
    }
    write_json(root / "manifest.json", manifest)


DB_RELATIVE_PATH = Path("enterprise") / "database" / "chainguard_enterprise_demo.db"


def _resolve_output_root(output_root: Path) -> Path:
    resolved = output_root.resolve()
    project_root = Path(__file__).resolve().parents[1]
    if resolved != project_root / "demo_assets" and project_root not in resolved.parents:
        raise ValueError(f"Output must stay inside project root: {project_root}")
    return resolved


def generate_db_only(output_root: Path) -> Path:
    """只重建 SQLite 库，其余资产一律不碰。

    干净检出（clone / worktree / CI）跑不了测试：demo_assets 下的 CSV/PDF/xlsx 都在
    版本库里，唯独 *.db 被 .gitignore 排除，于是 24 个用例以 scanned=0 失败——注意
    是空结果而不是报错，很容易被误判成代码回归。

    整体 generate() 能补上这个库，但它 rmtree 后重写全部资产，会把 60+ 个受控文件
    变成 modified（PDF 每次生成的字节都不同），代价远大于收益，还容易被误提交。
    数据生成是确定性的（固定 SEED 与 AS_OF、random.Random(SEED)），所以单独重建库
    与整体生成得到的库等价。
    """
    resolved = _resolve_output_root(output_root)
    db_path = resolved / DB_RELATIVE_PATH
    write_sqlite(db_path, build_data())
    return db_path


def generate(output_root: Path) -> None:
    resolved = _resolve_output_root(output_root)
    if resolved.exists():
        shutil.rmtree(resolved)
    resolved.mkdir(parents=True)

    data = build_data()
    csv_root = resolved / "enterprise" / "csv"
    for dataset_name, rows in data.items():
        write_csv(csv_root / f"{dataset_name}.csv", rows)
    write_sqlite(resolved / "enterprise" / "database" / "chainguard_enterprise_demo.db", data)
    write_json(
        resolved / "enterprise" / "json" / "dashboard_summary.json",
        {
            "company": COMPANY,
            "as_of": iso(AS_OF),
            "materials": len(data["materials"]),
            "inventory_records": len(data["inventory"]),
            "inventory_shortage_records": sum(
                row["available_qty"] < row["safety_stock_qty"] for row in data["inventory"]
            ),
            "at_risk_orders": sum(
                row["order_status"] in {"at_risk", "overdue"} for row in data["sales_orders"]
            ),
            "delayed_shipments": sum(row["delay_hours"] > 0 for row in data["shipments"]),
            "active_events": sum(
                row["event_status"] in {"active", "escalated"} for row in data["disruption_events"]
            ),
        },
    )
    write_small_business_xlsx(
        resolved / "small_business" / "ChainGuard小企业供应链演示账套.xlsx",
        data,
    )
    write_pdfs(resolved / "pdf", data)
    write_api_files(resolved / "erp_api")
    build_manifest(resolved, data)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "demo_assets",
        help="Output directory inside the project root.",
    )
    parser.add_argument(
        "--db-only",
        action="store_true",
        help=(
            "只重建 SQLite 库（干净检出跑测试所需），不重写 CSV/PDF/xlsx 等受控资产。"
        ),
    )
    args = parser.parse_args()
    if args.db_only:
        print(f"Rebuilt demo database at {generate_db_only(args.output)}")
        return
    generate(args.output)
    print(f"Generated demo assets at {args.output.resolve()}")


if __name__ == "__main__":
    main()
