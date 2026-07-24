from __future__ import annotations

import os
from dataclasses import asdict
from io import BytesIO
from pathlib import Path
import csv
import hashlib
import uuid
import zipfile
from typing import Annotated, Any

from fastapi import APIRouter, Depends, Query, UploadFile
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from src.security.encryption import encryption_status

from .. import sso as sso_service
from ..account_lifecycle import create_invitation, list_invitations, list_password_resets, lock_state, mask_account, resolve_pending_resets, revoke_invitation, unlock_account
from ..auth import AuthContext, can_view_data, get_current_user, require_permission
from ..config import settings
from ..database import get_db
from ..errors import ApiError
from ..models import CustomField, DataRecord, Department, ImportJob, Role, Tenant, User
from ..entity_import import DuplicateImportError, reserve_import_signature
from ..entity_repository import MODEL_BY_RESOURCE, list_product_rows, save_product_entity
from ..erp_integration import connector_for_config, get_config as get_erp_config, public_config as public_erp_config, record_test_result, safe_erp_error, save_config as save_erp_config
from ..erp_mapping_config import mapping_view as erp_mapping_view, reset_mapping as reset_erp_mapping, resolve_mapping as resolve_erp_mapping, review as review_erp_mapping, save_mapping as save_erp_mapping
from ..calibration_governance import build_governance_snapshot, confirm_governance_snapshot
from ..enterprise_import_catalog import IMPORT_TYPE_CATALOG, catalog_payload
from ..import_classifier import recognize_import_type
from ..jobs import enqueue_import_job, prepare_import_job
from ..onboarding import activate_tenant_after_business_data, inject_demo_dataset, onboarding_status, save_onboarding_progress
from ..org_settings import approval_chain_view, data_scope_view, save_approval_chain, save_data_scope
from ..reports import DEFAULT_MONTHS as REPORT_DEFAULT_MONTHS, executive_report as build_executive_report, operation_report as build_operation_report, response_report as build_response_report
from ..repository import add_audit, get_tenant_record, list_tenant_records, serialize
from ..notifications import ensure_rules, notify_event
from ..schemas import PatchRequest, TenantSettingsUpdate


router = APIRouter(tags=["imports-settings"])

STRUCTURED_SUFFIXES = {".csv", ".xlsx"}
OCR_SUFFIXES = {".pdf", ".doc", ".docx", ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
ARCHIVE_SUFFIXES = {".zip"}


def _headers_for_path(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(next(csv.reader(handle), []))
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        sheet = load_workbook(path, read_only=True, data_only=True).active
        return [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True), ())]
    return []


def _recognize_path(path: Path) -> dict[str, Any]:
    try:
        return recognize_import_type(path.name, _headers_for_path(path))
    except Exception:
        return recognize_import_type(path.name)


def _public_job(job: ImportJob) -> dict[str, Any]:
    payload = serialize(job)
    options = dict(payload.get("options") or {})
    options.pop("path", None)
    raw_result = dict(payload.get("result") or {})
    streaming = raw_result.get("streaming") if isinstance(raw_result.get("streaming"), dict) else {}

    def count(*keys: str) -> int:
        for source in (raw_result, streaming, payload):
            for key in keys:
                if source.get(key) is not None:
                    try:
                        return int(source[key])
                    except (TypeError, ValueError):
                        continue
        return 0

    source_rows = count("total", "sourceRows")
    success_rows = count("success", "successRows", "imported")
    rejected_rows = count("failed", "rejectedRows")
    reports = raw_result.get("reports") or raw_result.get("tableReports") or streaming.get("reports") or streaming.get("tableReports") or []
    normalized_result = {
        **raw_result,
        "total": source_rows,
        "sourceRows": source_rows,
        "success": success_rows,
        "successRows": success_rows,
        "failed": rejected_rows,
        "rejectedRows": rejected_rows,
        "reports": reports,
        "tableReports": reports,
    }
    operator = payload.get("operator") or options.get("operator") or "-"
    updated_at = job.updated_at.isoformat() if job.updated_at is not None else payload.get("createdAt")
    payload.update({
        "options": options,
        "result": normalized_result,
        "total": source_rows,
        "sourceRows": source_rows,
        "success": success_rows,
        "successRows": success_rows,
        "failed": rejected_rows,
        "rejectedRows": rejected_rows,
        "reports": reports,
        "operator": operator,
        "updatedAt": updated_at,
    })
    return payload


def _camel(name: str) -> str:
    parts = name.split("_")
    return parts[0] + "".join(part.title() for part in parts[1:])


def camelize_report(report: Any) -> dict[str, Any]:
    """P1-4：preflight 结果统一转 camelCase，与前端 estimatedRows/canProceed 契约对齐。"""
    data = asdict(report) if hasattr(report, "__dataclass_fields__") else dict(report)
    return {_camel(key): value for key, value in data.items()}


def normalize_xlsx_to_csv(source: str | Path) -> Path:
    """P1-4：XLSX 先解析归一化为 CSV，预检必须基于真实行而不是二进制字节估算。

    解析失败必须抛出异常（由调用方转红灯），不允许吞异常后继续绿灯。
    """
    from openpyxl import load_workbook

    path = Path(source)
    sheet = load_workbook(path, read_only=True, data_only=True).active
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values, ())]
    if not any(headers):
        raise ValueError("XLSX 首行没有可用表头")
    target = path.with_suffix(".csv")
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in values:
            writer.writerow(["" if value is None else value for value in row[: len(headers)]])
    return target


def normalized_preview(path: str | Path, limit: int = 20) -> dict[str, Any]:
    """Return a bounded, structured preview of the server-side normalized file."""
    source = Path(path)
    rows: list[dict[str, Any]] = []
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if len(rows) < limit:
                    rows.append(dict(row))
                else:
                    break
    elif source.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
            sheet = load_workbook(source, read_only=True, data_only=True).active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values, ())]
            for values_row in values:
                if len(rows) >= limit:
                    break
                rows.append({headers[index]: value for index, value in enumerate(values_row) if index < len(headers) and headers[index]})
        except Exception:
            rows = []
    return {"table": source.stem, "previewRows": rows, "previewLimit": limit}


@router.get("/imports/catalog")
def import_catalog(ctx: Annotated[AuthContext, Depends(require_permission("data:import"))]):
    return catalog_payload()


@router.post("/imports/upload", status_code=201)
async def upload_import(
    file: UploadFile,
    import_type: str = Query("auto", alias="type"),
    mode: str = Query("structured"),
    ctx: Annotated[AuthContext, Depends(require_permission("data:import"))] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    job_id = f"import-{uuid.uuid4().hex}"
    safe_name = Path(file.filename or "upload.csv").name
    suffix = Path(safe_name).suffix.lower()
    if mode not in {"structured", "ocr"}:
        raise ApiError(422, "CG-2603", "文件上传仅支持 structured 或 ocr 通道")
    allowed = STRUCTURED_SUFFIXES if mode == "structured" else OCR_SUFFIXES
    if suffix not in allowed:
        channel = "CSV/XLSX" if mode == "structured" else "PDF/Word/图片"
        raise ApiError(422, "CG-2603", f"当前通道仅支持 {channel}")
    if import_type != "auto" and import_type not in IMPORT_TYPE_CATALOG:
        raise ApiError(422, "CG-2603", "资料类型不存在")
    directory = Path(".workspace") / "imports" / ctx.tenant_id / job_id
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / safe_name
    size = 0
    with path.open("wb") as output:
        while chunk := await file.read(1024 * 1024):
            size += len(chunk)
            if size > settings.max_import_bytes:
                output.close()
                path.unlink(missing_ok=True)
                directory.rmdir()
                raise ApiError(413, "CG-2604", "上传文件超过大小限制")
            output.write(chunk)
    recognition = _recognize_path(path)
    job = ImportJob(
        id=job_id, tenant_id=ctx.tenant_id, file_name=safe_name, import_type=import_type,
        status="uploaded", progress=0,
        options={"size": size, "path": str(path.resolve()), "mode": mode, "recognition": recognition, "operator": ctx.name}, result={},
    )
    db.add(job); add_audit(db, ctx, "上传导入", "import", job.id, job.file_name, {"type": import_type, "mode": mode, "size": size}); db.commit()
    return _public_job(job)


@router.post("/imports/batch/classify", status_code=201)
async def classify_import_batch(
    files: list[UploadFile],
    ctx: Annotated[AuthContext, Depends(require_permission("data:import"))] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    """Upload a folder selection or ZIP and create one classified job per supported file."""

    candidates: list[tuple[str, bytes]] = []
    total_bytes = 0
    for uploaded in files:
        name = Path(uploaded.filename or "upload").name
        data = await uploaded.read()
        if Path(name).suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(BytesIO(data)) as archive:
                    members = [member for member in archive.infolist() if not member.is_dir()]
                    if len(members) > 200:
                        raise ApiError(413, "CG-2604", "压缩包文件数超过 200")
                    for member in members:
                        member_path = Path(member.filename.replace("\\", "/"))
                        if member_path.is_absolute() or ".." in member_path.parts:
                            raise ApiError(422, "CG-2606", "压缩包包含不安全路径")
                        if ((member.external_attr >> 16) & 0o170000) == 0o120000:
                            raise ApiError(422, "CG-2606", "压缩包不允许包含符号链接")
                        suffix = member_path.suffix.lower()
                        if suffix not in STRUCTURED_SUFFIXES | OCR_SUFFIXES:
                            continue
                        if member.file_size > settings.max_import_bytes:
                            raise ApiError(413, "CG-2604", f"{member_path.name} 超过单文件大小限制")
                        content = archive.read(member)
                        candidates.append((member_path.name, content))
                        total_bytes += len(content)
            except zipfile.BadZipFile as error:
                raise ApiError(422, "CG-2606", "ZIP 文件损坏或格式无效") from error
        else:
            suffix = Path(name).suffix.lower()
            if suffix not in STRUCTURED_SUFFIXES | OCR_SUFFIXES:
                continue
            candidates.append((name, data))
            total_bytes += len(data)
    if total_bytes > settings.max_import_bytes * 10:
        raise ApiError(413, "CG-2604", "批量上传总大小超过限制")
    if not candidates:
        raise ApiError(422, "CG-2603", "没有发现可导入的 CSV、XLSX、PDF、Word 或图片")

    batch_id = f"batch-{uuid.uuid4().hex}"
    jobs: list[dict[str, Any]] = []
    for index, (name, data) in enumerate(candidates, 1):
        job_id = f"import-{uuid.uuid4().hex}"
        directory = Path(".workspace") / "imports" / ctx.tenant_id / job_id
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / Path(name).name
        path.write_bytes(data)
        mode = "structured" if path.suffix.lower() in STRUCTURED_SUFFIXES else "ocr"
        recognition = _recognize_path(path)
        job = ImportJob(
            id=job_id, tenant_id=ctx.tenant_id, file_name=path.name, import_type="auto",
            status="uploaded", progress=0,
            options={"size": len(data), "path": str(path.resolve()), "mode": mode, "batchId": batch_id, "recognition": recognition, "operator": ctx.name},
            result={},
        )
        db.add(job)
        jobs.append({"jobId": job_id, "fileName": path.name, "mode": mode, "recognition": recognition, "sequence": index})
    add_audit(db, ctx, "批量上传并分类", "import", batch_id, batch_id, {"files": len(jobs), "bytes": total_bytes})
    db.commit()
    return {"batchId": batch_id, "files": jobs, "total": len(jobs), "requiresConfirmation": True}


def _erp_connector(values: dict[str, Any]):
    from src.connectors.rest_connector import RestErpConnector
    base_url = str(values.get("baseUrl") or "").strip()
    if not base_url.startswith(("http://", "https://")):
        raise ApiError(422, "CG-2610", "ERP 地址必须是 http:// 或 https:// URL")
    return RestErpConnector(base_url, api_key=str(values.get("apiKey") or ""), timeout=8.0, retries=1)


@router.post("/imports/erp/test")
def test_erp_connection(body: PatchRequest, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))]):
    try:
        return _erp_connector(body.values).test_connection()
    except Exception as error:
        raise ApiError(502, "CG-2611", safe_erp_error(error)) from error


@router.post("/imports/erp/preview")
def preview_erp_import(body: PatchRequest, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))]):
    connector = _erp_connector(body.values)
    selected = body.values.get("types") or list(IMPORT_TYPE_CATALOG)
    unknown = sorted(set(selected) - set(IMPORT_TYPE_CATALOG))
    if unknown:
        raise ApiError(422, "CG-2603", f"未知资料类型：{', '.join(unknown)}")
    resources = []
    try:
        for value in selected:
            definition = IMPORT_TYPE_CATALOG[value]
            rows = connector.fetch_resource(definition["erp_resource"])
            resources.append({"type": value, "label": definition["label"], "rows": len(rows), "preview": rows[:3]})
    except Exception as error:
        raise ApiError(502, "CG-2611", safe_erp_error(error)) from error
    return {"resources": resources, "totalRows": sum(item["rows"] for item in resources), "requiresConfirmation": True}


@router.post("/imports/erp/sync", status_code=201)
def sync_erp_import(
    body: PatchRequest,
    ctx: Annotated[AuthContext, Depends(require_permission("data:import"))],
    db: Annotated[Session, Depends(get_db)],
    connector: Any | None = None,
):
    """Run an explicitly confirmed, tenant-scoped ERP sync through the shared adapter."""

    if not bool(body.values.get("confirmed")):
        raise ApiError(409, "CG-2612", "必须确认 ERP 同步范围后才能执行")
    selected = list(dict.fromkeys(body.values.get("types") or []))
    if not selected or any(value not in IMPORT_TYPE_CATALOG for value in selected):
        raise ApiError(422, "CG-2603", "请选择有效的 ERP 资料类型")
    connector = connector or _erp_connector(body.values)
    from ..entity_import import aggregate_shipments, import_audit_rows, import_entity_rows

    # Resolve before any row is fetched: an unusable mapping must stop the sync,
    # never fall back to the shipped baseline behind the operator's back.
    spec, mapping_meta = resolve_erp_mapping(db, ctx.tenant_id)
    missing = [value for value in selected if IMPORT_TYPE_CATALOG[value]["entity"] and value not in spec["resources"]]
    if missing:
        raise ApiError(409, "CG-2810", f"当前 ERP 字段映射未声明资料类型：{', '.join(missing)}")

    order = [value for value in ("material", "supplier", "customer", "supplier_material", "order", "order_line", "inventory") if value in selected]
    order += [value for value in selected if value not in order]
    job_id = f"erp-{uuid.uuid4().hex}"
    job_options = {
        "mode": "erp",
        "baseUrl": str(getattr(connector, "base_url", body.values.get("baseUrl") or "")),
        "types": selected,
        "operator": ctx.name,
        "mappingSource": mapping_meta["source"],
        "mappingVersion": mapping_meta["version"],
        "mappingUpdatedAt": mapping_meta["updatedAt"],
        "mappingUpdatedBy": mapping_meta["updatedBy"],
    }
    job = ImportJob(
        id=job_id, tenant_id=ctx.tenant_id, file_name="ERP 接口同步", import_type="erp",
        status="running", progress=10, options=job_options, result={},
    )
    db.add(job)
    reports: list[dict[str, Any]] = []
    fetched: dict[str, list[dict[str, Any]]] = {}
    try:
        for value in order:
            definition = IMPORT_TYPE_CATALOG[value]
            rows = connector.fetch_resource(definition["erp_resource"])
            fetched[value] = rows
            if definition["entity"]:
                report = import_entity_rows(db, ctx.tenant_id, job_id, rows, value, spec=spec)
            else:
                report = import_audit_rows(db, ctx.tenant_id, job_id, rows, definition["source_table"])
            reports.append({"type": value, "label": definition["label"], **report})
        if "shipment" in fetched:
            po_lines = fetched.get("purchase_order_line") or connector.fetch_resource("purchase-order-lines")
            aggregate = aggregate_shipments(db, ctx.tenant_id, job_id, fetched["shipment"], po_lines)
            next(report for report in reports if report["type"] == "shipment")["aggregation"] = aggregate
        total = sum(int(report["sourceRows"]) for report in reports)
        rejected = sum(int(report["rejectedRows"]) for report in reports)
        job.status, job.progress = "succeeded", 100
        job.result = {"total": total, "success": total - rejected, "failed": rejected, "reports": reports}
        activate_tenant_after_business_data(db, ctx.tenant_id)
        add_audit(db, ctx, "ERP 接口导入", "import", job_id, "ERP 接口同步", {
            "types": selected, "total": total,
            "mappingSource": mapping_meta["source"], "mappingVersion": mapping_meta["version"],
        })
        ensure_rules(db, ctx.tenant_id)
        notify_event(db, ctx.tenant_id, "import_succeeded", {"trigger_user_id": ctx.user_id, "title": "ERP 同步完成", "target": "/settings/integration"})
        db.commit()
        return _public_job(job)
    except Exception as error:
        db.rollback()
        failed = ImportJob(
            id=job_id, tenant_id=ctx.tenant_id, file_name="ERP 接口同步", import_type="erp",
            status="failed", progress=100, options=job_options,
            result={"total": 0, "success": 0, "failed": 0, "reports": [], "errorSummary": safe_erp_error(error)},
        )
        db.add(failed)
        ensure_rules(db, ctx.tenant_id)
        notify_event(db, ctx.tenant_id, "import_failed", {"trigger_user_id": ctx.user_id, "title": "ERP 同步失败", "target": "/settings/integration"})
        db.commit()
        raise ApiError(502, "CG-2613", safe_erp_error(error)) from error


@router.get("/settings/encryption")
def encryption_settings(ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))]):
    """凭证静态加密的只读状态。

    加密是否可用是部署级事实（依赖库 + CHAINGUARD_ENCRYPTION_KEY），不是租户数据，
    因此这里不查库、不带 tenant 维度。此前它只在 Streamlit 演示（app.py 经
    src/security/posture.py）里可见，Web 端管理员看不到——凭证保存被拒时无从判断
    是部署没配密钥还是自己填错了。

    只回状态与派生方式，绝不回任何密钥材料：encryption_status() 本身就是只读探针，
    不做加解密、不抛异常，返回值里没有密钥或密文。
    """
    return encryption_status()


@router.get("/settings/integrations/erp")
def erp_integration_settings(
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    """Return only tenant-owned, credential-masked ERP integration state."""
    return public_erp_config(get_erp_config(db, ctx.tenant_id))


@router.patch("/settings/integrations/erp")
def save_erp_integration_settings(
    body: PatchRequest,
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    item = save_erp_config(db, ctx.tenant_id, body.values)
    add_audit(db, ctx, "保存 ERP 集成配置", "erp_integration", item.id, "ERP 集成", {"baseUrl": item.base_url, "credentialConfigured": bool(item.credential_ciphertext)})
    db.commit()
    return public_erp_config(item)


@router.post("/settings/integrations/erp/test")
def test_saved_erp_integration(
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    item = get_erp_config(db, ctx.tenant_id)
    if item is None:
        raise ApiError(409, "CG-2803", "请先保存 ERP 集成配置")
    try:
        result = connector_for_config(item).test_connection()
        record_test_result(item, result=result)
        db.commit()
        return {"ok": True, **public_erp_config(item)}
    except Exception as error:
        record_test_result(item, error=error)
        db.commit()
        raise ApiError(502, "CG-2804", item.last_test_error or safe_erp_error(error)) from error


@router.post("/settings/integrations/erp/sync", status_code=201)
def sync_saved_erp_integration(
    body: PatchRequest,
    ctx: Annotated[AuthContext, Depends(require_permission("data:import"))],
    db: Annotated[Session, Depends(get_db)],
):
    item = get_erp_config(db, ctx.tenant_id)
    if item is None:
        raise ApiError(409, "CG-2803", "请先保存 ERP 集成配置")
    values = {"confirmed": True, "types": body.values.get("types") or []}
    return sync_erp_import(PatchRequest(values=values), ctx, db, connector_for_config(item))


@router.get("/settings/integrations/erp/mapping")
def get_erp_mapping(
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    """Return the mapping this tenant's next ERP sync will use, with its provenance."""
    return erp_mapping_view(db, ctx.tenant_id)


@router.post("/settings/integrations/erp/mapping:validate")
def validate_erp_mapping(
    body: PatchRequest,
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
):
    """Dry-run a draft mapping; blocking errors and dangerous-but-allowed warnings are separate."""
    verdict = review_erp_mapping(body.values.get("spec") if isinstance(body.values.get("spec"), dict) else body.values)
    return {"valid": not verdict["errors"], **verdict}


@router.put("/settings/integrations/erp/mapping")
def put_erp_mapping(
    body: PatchRequest,
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    spec = body.values.get("spec") if isinstance(body.values.get("spec"), dict) else body.values
    meta = save_erp_mapping(db, ctx.tenant_id, spec, operator=ctx.name)
    add_audit(db, ctx, "保存 ERP 字段映射", "erp_field_mapping", f"{ctx.tenant_id}:{meta['version']}", "ERP 字段映射", {
        "version": meta["version"], "resources": sorted((spec.get("resources") or {}).keys()),
        "warnings": meta["warnings"],
    })
    db.commit()
    return {**erp_mapping_view(db, ctx.tenant_id), "warnings": meta["warnings"]}


@router.post("/settings/integrations/erp/mapping:reset")
def reset_erp_mapping_endpoint(
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    """Deactivate the tenant override so the shipped baseline applies again."""
    reset_erp_mapping(db, ctx.tenant_id)
    add_audit(db, ctx, "恢复 ERP 内置字段映射", "erp_field_mapping", ctx.tenant_id, "ERP 字段映射", {})
    db.commit()
    return erp_mapping_view(db, ctx.tenant_id)


@router.get("/settings/integrations/erp/mapping/source-fields")
def erp_mapping_source_fields(
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
    resource: Annotated[str, Query(min_length=1, max_length=60)],
):
    """Discover real source columns from ERP; every unavailable path names its reason."""
    if resource not in IMPORT_TYPE_CATALOG:
        raise ApiError(422, "CG-2603", "请选择有效的 ERP 资料类型")
    item = get_erp_config(db, ctx.tenant_id)
    if item is None:
        raise ApiError(409, "CG-2803", "请先保存 ERP 集成配置后再读取字段目录")
    if item.last_test_status != "available":
        raise ApiError(409, "CG-2813", "ERP 连接尚未通过测试，无法读取字段目录")
    try:
        rows = connector_for_config(item).sample_resource(IMPORT_TYPE_CATALOG[resource]["erp_resource"], limit=5)
    except Exception as error:
        raise ApiError(502, "CG-2814", safe_erp_error(error)) from error
    if not rows:
        raise ApiError(409, "CG-2814", "ERP 该资料类型没有可采样的数据，字段目录不可用")
    spec, _ = resolve_erp_mapping(db, ctx.tenant_id)
    rule = spec["resources"].get(resource) or {}
    mapped = {str(key) for key in (rule.get("fields") or {})}
    mapped |= {str(entry.get("from")) for entry in (rule.get("converts") or {}).values() if isinstance(entry, dict)}
    sensitive = {str(name).casefold() for name in spec.get("sensitive_columns") or []}
    fields: list[dict[str, Any]] = []
    for name in dict.fromkeys(key for row in rows for key in row):
        sample = next((row[name] for row in rows if row.get(name) not in (None, "")), None)
        fields.append({
            "name": str(name),
            "sample": None if str(name).casefold() in sensitive else str(sample)[:80] if sample is not None else None,
            "mapped": str(name) in mapped,
            "sensitive": str(name).casefold() in sensitive,
        })
    return {"resource": resource, "sampledRows": len(rows), "fields": fields}


@router.post("/imports/{item_id}/preflight")
def preflight(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    from src.import_preflight import run_preflight
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id)
    suffix = Path(item.file_name).suffix.lower()
    mode = str(item.options.get("mode") or ("ocr" if suffix in OCR_SUFFIXES else "structured"))
    if suffix in OCR_SUFFIXES:
        from src.ingestion_agent import ingest_files
        intake = ingest_files([item.options["path"]])
        extraction = intake.extractions[0]
        if extraction.needs_manual:
            item.status, item.progress = "manual_required", 25
            reason = extraction.note or "未提取到可校验内容。"
            item.result = {
                "canProceed": False, "status": "manual_required",
                "message": reason,
                "recognition": item.options.get("recognition"), "extraction": asdict(extraction),
                "manualReview": {
                    "required": True, "confirmationLevel": "full", "firstImport": True,
                    "reasonCode": extraction.error_code,
                    "suggestions": ["重新上传清晰、端正且保留表头和列分隔符的图片", "改用 CSV/Excel 上传，或人工录入"],
                },
            }
            db.commit(); payload = serialize(item); payload["options"].pop("path", None); return payload
        rows = next(iter(intake.normalized.values()), [])
        if rows and all(str(key).startswith("col_") for key in rows[0]):
            detected_headers = [str(value).strip() for value in rows[0].values()]
            if len(rows) > 1 and all(detected_headers):
                rows = [
                    {detected_headers[index]: value for index, value in enumerate(source.values()) if index < len(detected_headers)}
                    for source in rows[1:]
                ]
        if not rows:
            item.status, item.progress = "manual_required", 25
            item.result = {
                "canProceed": False, "status": "manual_required",
                "message": "已识别文档，但没有形成可导入的数据行；请人工处理原文。",
                "recognition": item.options.get("recognition"), "extraction": asdict(extraction),
                "manualReview": {"required": True, "confirmationLevel": "full", "firstImport": True},
            }
            db.commit(); return _public_job(item)
        normalized_path = Path(item.options["path"]).parent / "normalized.csv"
        with normalized_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader(); writer.writerows(rows)
        recognition = recognize_import_type(item.file_name, rows[0].keys() if rows else [])
        schema_signature = hashlib.sha256("|".join(sorted(rows[0].keys() if rows else [])).encode("utf-8")).hexdigest()
        previous = [
            job for job in db.query(ImportJob).filter(ImportJob.tenant_id == ctx.tenant_id, ImportJob.status == "succeeded").all()
            if job.options.get("mode") == "ocr" and job.options.get("schemaSignature") == schema_signature
        ]
        seen_count = len(previous)
        item.options = {
            **item.options, "originalPath": item.options["path"], "path": str(normalized_path.resolve()),
            "extraction": asdict(extraction), "recognition": recognition, "schemaSignature": schema_signature,
        }
    elif suffix == ".xlsx":
        # P1-4：XLSX 先归一化为 CSV 再做基于行的容量预检；解析失败必须红灯，禁止吞异常假绿灯
        try:
            normalized_path = normalize_xlsx_to_csv(item.options["path"])
        except Exception:
            item.status, item.progress = "failed", 25
            item.result = {
                "canProceed": False,
                "verdict": "PARSE_ERROR",
                "estimatedRows": None,
                "messages": ["XLSX 解析失败：文件可能损坏或不是有效的 Excel 工作簿，导入已阻止。"],
                "normalized": {"table": Path(item.file_name).stem, "previewRows": [], "previewLimit": 20},
            }
            db.commit(); payload = serialize(item); payload["options"].pop("path", None); return payload
        item.options = {**item.options, "originalPath": item.options["path"], "path": str(normalized_path.resolve())}
    report = run_preflight([item.options["path"]], Path(item.options["path"]).parent / "import.db")
    item.status = "manual_review" if mode == "ocr" and report.can_proceed else "preflighted" if report.can_proceed else "failed"
    item.progress = 25
    # camelCase 与前端 estimatedRows/canProceed 契约对齐（P1-4）；verdict 键名不变，confirm 闸门不受影响
    item.result = {
        **camelize_report(report), "normalized": normalized_preview(item.options["path"]),
        "recognition": item.options.get("recognition"),
    }
    if mode == "ocr":
        # Successful OCR must expose auditable engine/confidence/timing metadata
        # to the review UI; the server path remains private in job options.
        item.result["extraction"] = item.options.get("extraction")
        item.result["manualReview"] = {
            "required": True, "firstImport": seen_count == 0, "seenCount": seen_count,
            "familiarity": "novel" if seen_count == 0 else "familiar",
            "confirmationLevel": "full" if seen_count == 0 else "light",
            "confirmationPoints": ["识别出的资料类型", "关键业务字段", "低置信度/空值行"] if seen_count == 0 else ["本次差异与低置信度行"],
        }
    db.commit()
    payload = serialize(item); payload["options"].pop("path", None); return payload


@router.post("/imports/{item_id}/confirm")
def confirm_import(item_id: str, body: PatchRequest, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id)
    force = bool(body.values.get("force"))
    confirmed_type = str(body.values.get("confirmedType") or item.import_type)
    if item.import_type == "auto" and confirmed_type not in IMPORT_TYPE_CATALOG:
        raise ApiError(409, "CG-2607", "必须人工确认识别出的资料类型后才能继续")
    if confirmed_type not in IMPORT_TYPE_CATALOG:
        raise ApiError(422, "CG-2603", "资料类型不存在")
    mode = str(item.options.get("mode") or "structured")
    if mode == "ocr" and not bool(body.values.get("manualConfirmed")):
        raise ApiError(409, "CG-2608", "OCR/文档导入必须完成人工确认，不能强制绕过")
    if item.status == "failed":
        # Disk shortage is a hard safety gate, not an override-able quality warning.
        if item.result.get("verdict") == "INSUFFICIENT_DISK":
            raise ApiError(409, "CG-2602", "磁盘空间不足，不能强制导入")
        # P1-4：解析失败没有可导入的归一化数据，同样不允许强制
        if item.result.get("verdict") == "PARSE_ERROR":
            raise ApiError(409, "CG-2602", "文件解析失败，不能强制导入")
        if not force:
            raise ApiError(409, "CG-2602", "预检未通过，需明确确认后才能继续导入")
    if item.status not in {"preflighted", "manual_review", "failed"}:
        raise ApiError(409, "CG-2601", "导入任务尚未完成预检")
    item.import_type = confirmed_type
    item.status = "confirmed"; item.options = {**item.options, **body.values, "typeConfirmed": True}; item.progress = 50; db.commit()
    payload = serialize(item); payload["options"].pop("path", None); return payload


@router.post("/imports/{item_id}/execute", status_code=202)
def execute_import(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id)
    if item.status != "confirmed":
        raise ApiError(409, "CG-2601", "导入任务尚未确认")
    try:
        history = reserve_import_signature(
            db,
            ctx.tenant_id,
            item.id,
            item.file_name,
            item.options["path"],
            item.import_type,
        )
        db.flush()  # UNIQUE(tenant_id, signature) closes concurrent execute races.
    except DuplicateImportError as error:
        raise ApiError(409, "CG-2605", f"D04：相同签名文件已导入（批次 {error.history.import_job_id}）") from error
    except IntegrityError as error:
        db.rollback()
        raise ApiError(409, "CG-2605", "D04：相同签名文件已被其他导入任务占用") from error
    item.status = "pending"
    item.options = {**item.options, "signature": history.signature}
    prepare_import_job(db, item, ctx)
    db.commit()
    enqueue_import_job(item.id, ctx)
    return {"jobId": item.id, "status": item.status}


@router.get("/imports/{item_id}")
def import_progress(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id); return _public_job(item)
@router.get("/imports")
def import_history(ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    data = [_public_job(item) for item in list_tenant_records(db, ImportJob, ctx.tenant_id)]
    return {"data": data, "total": len(data), "success": True}
@router.post("/imports/{item_id}/rollback")
def rollback_import(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]): item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id); item.status = "rolled_back"; db.commit(); return {"ok": True, "id": item_id}


@router.get("/settings/users")
def users(ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    items = list_tenant_records(db, User, ctx.tenant_id)
    # 登录标识按 5B 账户完善要求脱敏后回显（字段保留，值不再是可直接撞库的原文）；
    # SSO subject 是 IdP 侧标识，不出接口。锁定状态随行返回，供"解锁"操作判断。
    data = [{**serialize(x, exclude={"sso_subject"}), "account": mask_account(x.account), "roleIds": [x.role_id], "ssoLinked": bool(x.sso_subject), **lock_state(x)} for x in items]
    return {"data": data, "total": len(data), "success": True}


@router.post("/settings/users", status_code=201)
def create_user(body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    from ..auth.security import hash_password
    role = get_tenant_record(db, Role, body["roleId"], ctx.tenant_id)
    item = User(id=f"u-{uuid.uuid4().hex}", tenant_id=ctx.tenant_id, account=str(body.get("account") or body.get("phone") or body.get("email")).lower(), password_hash=hash_password(str(body.get("password") or uuid.uuid4().hex)), name=body["name"], phone=body.get("phone", ""), email=body.get("email", ""), dept_id=body.get("deptId", "dept-1"), role_id=role.id, role_code=role.code, status=body.get("status", "active"), data_scope=body.get("dataScope", "custom")); db.add(item); add_audit(db, ctx, "创建用户", "user", item.id, item.name, {"roleCode": role.code}); db.commit(); return {"ok": True, "id": item.id}


@router.patch("/settings/users/{item_id}")
def update_user(item_id: str, body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, User, item_id, ctx.tenant_id)
    before = {"name": item.name, "status": item.status, "roleCode": item.role_code}
    for field in ("name", "phone", "email", "status"):
        if field in body: setattr(item, field, body[field])
    if "dataScope" in body: item.data_scope = body["dataScope"]
    if "roleId" in body:
        role = get_tenant_record(db, Role, body["roleId"], ctx.tenant_id); item.role_id = role.id; item.role_code = role.code
    add_audit(db, ctx, "更新用户", "user", item.id, item.name, {"before": before, "after": body}); db.commit(); return {"ok": True, "id": item.id}


@router.post("/settings/users/{item_id}/reset-password")
def reset_user_password(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    from ..auth.security import hash_password, revoke_refresh_tokens
    import secrets
    item = get_tenant_record(db, User, item_id, ctx.tenant_id)
    temporary_password = f"Cg!{secrets.token_urlsafe(9)}"
    item.password_hash, item.must_change_password = hash_password(temporary_password), True
    revoke_refresh_tokens(db, item)
    # 管理员兜底重置同时解锁并关掉该用户的找回密码待办，闭合"通道未配置"降级链路
    from ..account_lifecycle import clear_login_failures
    clear_login_failures(item)
    resolved = resolve_pending_resets(db, ctx, item)
    add_audit(db, ctx, "重置密码", "user", item.id, item.name, {"mustChangePassword": True, "resolvedResetRequests": resolved})
    db.commit()
    return {"ok": True, "temporaryPassword": temporary_password, "mustChangePassword": True, "resolvedResetRequests": resolved}


@router.post("/settings/users/{item_id}/unlock")
def unlock_user(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, User, item_id, ctx.tenant_id)
    result = unlock_account(db, ctx, item)
    db.commit()
    return result


@router.get("/settings/password-resets")
def password_resets(ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    data = list_password_resets(db, ctx.tenant_id)
    return {"data": data, "total": len(data), "success": True}


@router.get("/settings/invitations")
def invitations(ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    data = list_invitations(db, ctx.tenant_id)
    return {"data": data, "total": len(data), "success": True}


@router.post("/settings/invitations", status_code=201)
def add_invitation(body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    """返回体里的 code 是这枚邀请码明文唯一一次出现；库内与后续列表只有哈希与掩码。"""
    return create_invitation(db, ctx, body)


@router.post("/settings/invitations/{item_id}/revoke")
def disable_invitation(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    return revoke_invitation(db, ctx, item_id)


@router.get("/settings/sso")
def sso_config(ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    return sso_service.public_config(sso_service.get_config(db, ctx.tenant_id))


@router.put("/settings/sso")
def save_sso_config(body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    return sso_service.save_config(db, ctx, body)


@router.delete("/settings/users/{item_id}", status_code=204)
def disable_user(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, User, item_id, ctx.tenant_id); item.status = "disabled"; add_audit(db, ctx, "停用用户", "user", item.id, item.name, {}); db.commit()


@router.get("/settings/roles")
def roles(ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]): return [serialize(x) for x in list_tenant_records(db, Role, ctx.tenant_id)]
@router.post("/settings/roles", status_code=201)
def save_role(body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    item = Role(id=body.get("id") or f"role-{uuid.uuid4().hex}", tenant_id=ctx.tenant_id, code=body["code"], name=body["name"], builtin=False, permissions=body.get("permissions", [])); db.merge(item); add_audit(db, ctx, "保存角色", "role", item.id, item.name, {"permissions": item.permissions}); db.commit(); return {"ok": True, "id": item.id}
@router.patch("/settings/roles/{item_id}")
def update_role(item_id: str, body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, Role, item_id, ctx.tenant_id)
    if item.builtin:
        from ..errors import ApiError
        raise ApiError(409, "CG-2701", "内置角色不可修改，请复制后编辑")
    if "name" in body: item.name = body["name"]
    if "permissions" in body: item.permissions = body["permissions"]
    add_audit(db, ctx, "更新角色", "role", item.id, item.name, body); db.commit(); return {"ok": True, "id": item.id}
@router.delete("/settings/roles/{item_id}", status_code=204)
def delete_role(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]): item = get_tenant_record(db, Role, item_id, ctx.tenant_id); db.delete(item); db.commit()


@router.get("/settings/tenant")
def tenant(ctx: Annotated[AuthContext, Depends(get_current_user)], db: Annotated[Session, Depends(get_db)]): return serialize(db.get(Tenant, ctx.tenant_id))


@router.patch("/settings/tenant")
def update_tenant(
    body: TenantSettingsUpdate,
    ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    item = db.get(Tenant, ctx.tenant_id)
    if item is None:
        raise ApiError(404, "CG-2805", "租户不存在")
    changes = body.model_dump(exclude_none=True)
    for field, value in changes.items():
        value = value.strip()
        if not value:
            raise ApiError(422, "CG-2806", f"{field} 不能为空")
        setattr(item, field, value)
    if changes:
        add_audit(db, ctx, "更新企业信息", "tenant", item.id, item.name, changes)
        db.commit()
        db.refresh(item)
    return serialize(item)
@router.get("/settings/departments")
def departments(
    ctx: Annotated[AuthContext, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    """真实部门树（此前是端点里的 5 个硬编码字符串，没有层级也不可配置）。"""
    rows = list(db.scalars(select(Department).where(Department.tenant_id == ctx.tenant_id).order_by(Department.code)).all())
    return [
        {"id": row.id, "tenantId": row.tenant_id, "code": row.code, "name": row.name, "parentId": row.parent_id}
        for row in rows
    ]


@router.get("/settings/calibration-governance")
def calibration_governance(
    ctx: Annotated[AuthContext, Depends(require_permission("settings:approval"))],
    db: Annotated[Session, Depends(get_db)],
):
    """Read a tenant's existing-engine calibration proposal; no config writes."""
    snapshot = build_governance_snapshot(db, ctx.tenant_id)
    db.commit()  # persists only D3 notification messages/rules when drift exceeds limits
    return snapshot


@router.post("/settings/calibration-governance/confirm")
def confirm_calibration_governance(
    body: PatchRequest,
    ctx: Annotated[AuthContext, Depends(require_permission("settings:approval"))],
    db: Annotated[Session, Depends(get_db)],
):
    recommendation_id = str(body.values.get("recommendationId") or "")
    if not recommendation_id:
        raise ApiError(422, "CG-2901", "必须指定当前校准建议")
    try:
        result = confirm_governance_snapshot(db, ctx.tenant_id, ctx.user_id, recommendation_id)
    except ValueError as error:
        raise ApiError(409, "CG-2902", str(error)) from error
    add_audit(db, ctx, "确认校准建议", "calibration_governance", recommendation_id, "风险阈值与权重校准", result)
    db.commit()
    return {"ok": True, **result}


@router.get("/settings/custom-fields")
def fields(object_type: str = Query(..., alias="objectType"), ctx: Annotated[AuthContext, Depends(get_current_user)] = None, db: Annotated[Session, Depends(get_db)] = None): return [serialize(x) for x in list_tenant_records(db, CustomField, ctx.tenant_id) if x.object_type == object_type]
@router.post("/settings/custom-fields", status_code=201)
def save_field(body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]): item = CustomField(id=body.get("id") or f"field-{uuid.uuid4().hex}", tenant_id=ctx.tenant_id, object_type=body["objectType"], name=body["name"], label=body["label"], field_type=body.get("type", "string"), required=body.get("required", False), enabled=True, config=body.get("config", {})); db.merge(item); db.commit(); return {"ok": True, "id": item.id}
@router.delete("/settings/custom-fields/{item_id}")
def disable_field(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))], db: Annotated[Session, Depends(get_db)]): item = get_tenant_record(db, CustomField, item_id, ctx.tenant_id); item.enabled = False; db.commit(); return {"ok": True, "id": item.id}


@router.get("/data/{resource_type}")
def data_table(resource_type: str, ctx: Annotated[AuthContext, Depends(get_current_user)], db: Annotated[Session, Depends(get_db)]):
    if resource_type not in MODEL_BY_RESOURCE:
        raise ApiError(404, "CG-2804", "资料类型不存在")
    if not can_view_data(ctx, resource_type):
        raise ApiError(403, "CG-1003", "没有操作权限")
    items = list_product_rows(db, ctx.tenant_id, resource_type)
    return {"data": items, "total": len(items), "success": True}


@router.post("/data/{resource_type}", status_code=201)
def create_data_record(resource_type: str, body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("data:manage"))], db: Annotated[Session, Depends(get_db)]):
    if resource_type not in MODEL_BY_RESOURCE:
        raise ApiError(404, "CG-2804", "资料类型不存在")
    name = str(body.get("name") or "").strip()
    if not name and resource_type in {"material", "supplier", "customer"}:
        raise ApiError(422, "CG-2801", "名称不能为空")
    try:
        item = save_product_entity(db, ctx.tenant_id, resource_type, body)
    except ValueError as error:
        raise ApiError(422, "CG-2802", str(error)) from error
    rows = list_product_rows(db, ctx.tenant_id, resource_type)
    business_id = getattr(item, {"material": "material_id", "supplier": "supplier_id", "customer": "customer_id", "order": "sales_order_id", "inventory": "inventory_id"}[resource_type])
    payload = next(row for row in rows if row["id"] == business_id)
    add_audit(db, ctx, "新建资料", resource_type, str(business_id), str(payload.get("name") or payload.get("orderNo") or payload.get("material") or business_id), payload)
    db.commit()
    return payload


@router.get("/data/{resource_type}/{item_id}")
def data_record_detail(item_id: str, resource_type: str, ctx: Annotated[AuthContext, Depends(get_current_user)], db: Annotated[Session, Depends(get_db)]):
    if resource_type not in MODEL_BY_RESOURCE:
        raise ApiError(404, "CG-2804", "资料类型不存在")
    if not can_view_data(ctx, resource_type):
        raise ApiError(403, "CG-1003", "没有操作权限")
    item = next((row for row in list_product_rows(db, ctx.tenant_id, resource_type) if row["id"] == item_id), None)
    if item is None:
        raise ApiError(404, "CG-2804", "资料不存在")
    return item


@router.patch("/data/{resource_type}/{item_id}")
def update_data_record(item_id: str, resource_type: str, body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("data:manage"))], db: Annotated[Session, Depends(get_db)]):
    if resource_type not in MODEL_BY_RESOURCE:
        raise ApiError(404, "CG-2804", "资料类型不存在")
    try:
        save_product_entity(db, ctx.tenant_id, resource_type, body, business_key=item_id)
    except LookupError as error:
        raise ApiError(404, "CG-2804", "资料不存在") from error
    except ValueError as error:
        raise ApiError(422, "CG-2802", str(error)) from error
    payload = next(row for row in list_product_rows(db, ctx.tenant_id, resource_type) if row["id"] == item_id)
    add_audit(db, ctx, "更新资料", resource_type, item_id, str(payload.get("name") or payload.get("orderNo") or payload.get("material") or item_id), body)
    db.commit()
    return payload


@router.get("/risk-rules")
def risk_rules(ctx: Annotated[AuthContext, Depends(get_current_user)], db: Annotated[Session, Depends(get_db)]):
    items = [x for x in list_tenant_records(db, DataRecord, ctx.tenant_id) if x.resource_type == "risk_rule"]
    if not items:
        return {"data": [{"id": "rule-1", "name": "安全库存预警线", "threshold": "20%", "enabled": True}]}
    return {"data": [{"id": x.id, "name": x.name, **x.payload} for x in items]}


@router.put("/risk-rules/{item_id}")
def update_risk_rule(item_id: str, body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("risk:manage"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, DataRecord, item_id, ctx.tenant_id); item.payload = {**item.payload, **body}; add_audit(db, ctx, "更新风险规则", "risk_rule", item.id, item.name, body); db.commit(); return {"ok": True, "id": item.id, **item.payload}


@router.get("/notifications/webhook-config")
def webhook_config(ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))]): return {"enabled": bool(os.getenv("WEBHOOK_URL")), "url": "***" if os.getenv("WEBHOOK_URL") else ""}
@router.put("/notifications/webhook-config")
def update_webhook_config(body: dict[str, Any], ctx: Annotated[AuthContext, Depends(require_permission("settings:manage"))]): return {"enabled": bool(body.get("enabled")), "url": "***" if body.get("url") else ""}


@router.get("/reports/executive")
def executive_report(
    ctx: Annotated[AuthContext, Depends(require_permission("report:executive"))],
    db: Annotated[Session, Depends(get_db)],
    months: int = Query(REPORT_DEFAULT_MONTHS, ge=1, le=36),
):
    return build_executive_report(db, ctx.tenant_id, months)


@router.get("/reports/operation")
def operation_report(
    ctx: Annotated[AuthContext, Depends(require_permission("report:operation"))],
    db: Annotated[Session, Depends(get_db)],
    months: int = Query(REPORT_DEFAULT_MONTHS, ge=1, le=36),
):
    return build_operation_report(db, ctx.tenant_id, months)


@router.get("/reports/response")
def response_report(
    ctx: Annotated[AuthContext, Depends(require_permission("report:view"))],
    db: Annotated[Session, Depends(get_db)],
    months: int = Query(REPORT_DEFAULT_MONTHS, ge=1, le=36),
):
    return build_response_report(db, ctx.tenant_id, months)


@router.get("/settings/approval-chain")
def get_approval_chain(
    ctx: Annotated[AuthContext, Depends(require_permission("settings:approval"))],
    db: Annotated[Session, Depends(get_db)],
):
    return approval_chain_view(db, ctx.tenant_id)


@router.put("/settings/approval-chain")
def put_approval_chain(
    body: dict[str, Any],
    ctx: Annotated[AuthContext, Depends(require_permission("settings:approval"))],
    db: Annotated[Session, Depends(get_db)],
):
    result = save_approval_chain(db, ctx.tenant_id, body, actor=ctx.user_id)
    add_audit(db, ctx, "更新审批流", "approval_chain", "approval_chain", "审批链配置", result)
    db.commit()
    return result


@router.get("/settings/data-scopes")
def get_data_scopes(
    ctx: Annotated[AuthContext, Depends(require_permission("role:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    return data_scope_view(db, ctx.tenant_id)


@router.put("/settings/data-scopes")
def put_data_scopes(
    body: dict[str, Any],
    ctx: Annotated[AuthContext, Depends(require_permission("role:manage"))],
    db: Annotated[Session, Depends(get_db)],
):
    result = save_data_scope(db, ctx.tenant_id, body, actor=ctx.user_id)
    add_audit(db, ctx, "更新数据范围", "data_scope", "data_scope", "角色数据范围", body)
    db.commit()
    return result
@router.get("/onboarding/templates")
def templates(): return [{"id": "electronics", "name": "电子制造", "desc": "芯片、PCB、关键物料齐套与替代供应商模板"}]


@router.get("/onboarding/status")
def onboarding_status_endpoint(
    ctx: Annotated[AuthContext, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    """C3 status is always recomputed from this tenant's persisted C2 entities."""
    return onboarding_status(db, ctx.tenant_id)


@router.post("/onboarding/progress")
def save_progress(
    body: dict[str, Any],
    ctx: Annotated[AuthContext, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    status = save_onboarding_progress(db, ctx, body)
    db.commit()
    return {"ok": True, "status": status}


@router.post("/onboarding/demo-dataset", status_code=201)
def inject_onboarding_demo_dataset(
    body: PatchRequest,
    ctx: Annotated[AuthContext, Depends(require_permission("data:import"))],
    db: Annotated[Session, Depends(get_db)],
):
    # No implicit path: both UI and API callers must record an explicit second confirmation.
    if body.values.get("confirmed") is not True:
        raise ApiError(422, "CG-2701", "请确认后再注入演示数据集")
    try:
        result = inject_demo_dataset(db, ctx)
    except ValueError as error:
        raise ApiError(409, "CG-2702", str(error)) from error
    db.commit()
    return result


@router.post("/onboarding/templates/{template_id}/apply")
def apply_template(template_id: str, ctx: Annotated[AuthContext, Depends(get_current_user)], db: Annotated[Session, Depends(get_db)]): item = db.get(Tenant, ctx.tenant_id); item.industry = template_id; db.commit(); return {"ok": True, "tenant": serialize(item)}
