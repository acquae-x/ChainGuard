from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
from pathlib import Path
import csv
import uuid
import zipfile
from typing import Annotated, Any

from fastapi import APIRouter, Depends, Query, UploadFile
from sqlalchemy.orm import Session


from ..auth import AuthContext, require_permission
from ..database import get_db
from ..errors import ApiError
from ..models import ImportJob
from ..enterprise_import_catalog import IMPORT_TYPE_CATALOG, catalog_payload
from ..import_classifier import recognize_import_type
from ..repository import add_audit, serialize


router = APIRouter(tags=["imports-settings"])

STRUCTURED_SUFFIXES = {".csv", ".xlsx"}
OCR_SUFFIXES = {".pdf", ".doc", ".docx", ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
ARCHIVE_SUFFIXES = {".zip"}


def _max_import_bytes() -> int:
    from . import imports_settings

    return imports_settings.settings.max_import_bytes


def _headers_for_path(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(next(csv.reader(handle), []))
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        sheet = load_workbook(path, read_only=True, data_only=True).active
        return [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True), ())]
    return []


def _recognize_path(path: Path) -> dict[str, Any]:
    try:
        return recognize_import_type(path.name, _headers_for_path(path))
    except Exception:
        return recognize_import_type(path.name)


def _public_job(job: ImportJob) -> dict[str, Any]:
    payload = serialize(job)
    options = dict(payload.get("options") or {})
    options.pop("path", None)
    raw_result = dict(payload.get("result") or {})
    streaming = raw_result.get("streaming") if isinstance(raw_result.get("streaming"), dict) else {}

    def count(*keys: str) -> int:
        for source in (raw_result, streaming, payload):
            for key in keys:
                if source.get(key) is not None:
                    try:
                        return int(source[key])
                    except (TypeError, ValueError):
                        continue
        return 0

    source_rows = count("total", "sourceRows")
    success_rows = count("success", "successRows", "imported")
    rejected_rows = count("failed", "rejectedRows")
    reports = raw_result.get("reports") or raw_result.get("tableReports") or streaming.get("reports") or streaming.get("tableReports") or []
    normalized_result = {
        **raw_result,
        "total": source_rows,
        "sourceRows": source_rows,
        "success": success_rows,
        "successRows": success_rows,
        "failed": rejected_rows,
        "rejectedRows": rejected_rows,
        "reports": reports,
        "tableReports": reports,
    }
    operator = payload.get("operator") or options.get("operator") or "-"
    updated_at = job.updated_at.isoformat() if job.updated_at is not None else payload.get("createdAt")
    payload.update({
        "options": options,
        "result": normalized_result,
        "total": source_rows,
        "sourceRows": source_rows,
        "success": success_rows,
        "successRows": success_rows,
        "failed": rejected_rows,
        "rejectedRows": rejected_rows,
        "reports": reports,
        "operator": operator,
        "updatedAt": updated_at,
    })
    return payload


def _camel(name: str) -> str:
    parts = name.split("_")
    return parts[0] + "".join(part.title() for part in parts[1:])


def camelize_report(report: Any) -> dict[str, Any]:
    """P1-4：preflight 结果统一转 camelCase，与前端 estimatedRows/canProceed 契约对齐。"""
    data = asdict(report) if hasattr(report, "__dataclass_fields__") else dict(report)
    return {_camel(key): value for key, value in data.items()}


def normalize_xlsx_to_csv(source: str | Path) -> Path:
    """P1-4：XLSX 先解析归一化为 CSV，预检必须基于真实行而不是二进制字节估算。

    解析失败必须抛出异常（由调用方转红灯），不允许吞异常后继续绿灯。
    """
    from openpyxl import load_workbook

    path = Path(source)
    sheet = load_workbook(path, read_only=True, data_only=True).active
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values, ())]
    if not any(headers):
        raise ValueError("XLSX 首行没有可用表头")
    target = path.with_suffix(".csv")
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in values:
            writer.writerow(["" if value is None else value for value in row[: len(headers)]])
    return target


def normalized_preview(path: str | Path, limit: int = 20) -> dict[str, Any]:
    """Return a bounded, structured preview of the server-side normalized file."""
    source = Path(path)
    rows: list[dict[str, Any]] = []
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if len(rows) < limit:
                    rows.append(dict(row))
                else:
                    break
    elif source.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
            sheet = load_workbook(source, read_only=True, data_only=True).active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values, ())]
            for values_row in values:
                if len(rows) >= limit:
                    break
                rows.append({headers[index]: value for index, value in enumerate(values_row) if index < len(headers) and headers[index]})
        except Exception:
            rows = []
    return {"table": source.stem, "previewRows": rows, "previewLimit": limit}


@router.get("/imports/catalog")
def import_catalog(ctx: Annotated[AuthContext, Depends(require_permission("data:import"))]):
    return catalog_payload()


@router.post("/imports/upload", status_code=201)
async def upload_import(
    file: UploadFile,
    import_type: str = Query("auto", alias="type"),
    mode: str = Query("structured"),
    ctx: Annotated[AuthContext, Depends(require_permission("data:import"))] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    job_id = f"import-{uuid.uuid4().hex}"
    safe_name = Path(file.filename or "upload.csv").name
    suffix = Path(safe_name).suffix.lower()
    if mode not in {"structured", "ocr"}:
        raise ApiError(422, "CG-2603", "文件上传仅支持 structured 或 ocr 通道")
    allowed = STRUCTURED_SUFFIXES if mode == "structured" else OCR_SUFFIXES
    if suffix not in allowed:
        channel = "CSV/XLSX" if mode == "structured" else "PDF/Word/图片"
        raise ApiError(422, "CG-2603", f"当前通道仅支持 {channel}")
    if import_type != "auto" and import_type not in IMPORT_TYPE_CATALOG:
        raise ApiError(422, "CG-2603", "资料类型不存在")
    directory = Path(".workspace") / "imports" / ctx.tenant_id / job_id
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / safe_name
    size = 0
    with path.open("wb") as output:
        while chunk := await file.read(1024 * 1024):
            size += len(chunk)
            if size > _max_import_bytes():
                output.close()
                path.unlink(missing_ok=True)
                directory.rmdir()
                raise ApiError(413, "CG-2604", "上传文件超过大小限制")
            output.write(chunk)
    recognition = _recognize_path(path)
    job = ImportJob(
        id=job_id, tenant_id=ctx.tenant_id, file_name=safe_name, import_type=import_type,
        status="uploaded", progress=0,
        options={"size": size, "path": str(path.resolve()), "mode": mode, "recognition": recognition, "operator": ctx.name}, result={},
    )
    db.add(job); add_audit(db, ctx, "上传导入", "import", job.id, job.file_name, {"type": import_type, "mode": mode, "size": size}); db.commit()
    return _public_job(job)


@router.post("/imports/batch/classify", status_code=201)
async def classify_import_batch(
    files: list[UploadFile],
    ctx: Annotated[AuthContext, Depends(require_permission("data:import"))] = None,
    db: Annotated[Session, Depends(get_db)] = None,
):
    """Upload a folder selection or ZIP and create one classified job per supported file."""

    candidates: list[tuple[str, bytes]] = []
    total_bytes = 0
    for uploaded in files:
        name = Path(uploaded.filename or "upload").name
        data = await uploaded.read()
        if Path(name).suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(BytesIO(data)) as archive:
                    members = [member for member in archive.infolist() if not member.is_dir()]
                    if len(members) > 200:
                        raise ApiError(413, "CG-2604", "压缩包文件数超过 200")
                    for member in members:
                        member_path = Path(member.filename.replace("\\", "/"))
                        if member_path.is_absolute() or ".." in member_path.parts:
                            raise ApiError(422, "CG-2606", "压缩包包含不安全路径")
                        if ((member.external_attr >> 16) & 0o170000) == 0o120000:
                            raise ApiError(422, "CG-2606", "压缩包不允许包含符号链接")
                        suffix = member_path.suffix.lower()
                        if suffix not in STRUCTURED_SUFFIXES | OCR_SUFFIXES:
                            continue
                        if member.file_size > _max_import_bytes():
                            raise ApiError(413, "CG-2604", f"{member_path.name} 超过单文件大小限制")
                        content = archive.read(member)
                        candidates.append((member_path.name, content))
                        total_bytes += len(content)
            except zipfile.BadZipFile as error:
                raise ApiError(422, "CG-2606", "ZIP 文件损坏或格式无效") from error
        else:
            suffix = Path(name).suffix.lower()
            if suffix not in STRUCTURED_SUFFIXES | OCR_SUFFIXES:
                continue
            candidates.append((name, data))
            total_bytes += len(data)
    if total_bytes > _max_import_bytes() * 10:
        raise ApiError(413, "CG-2604", "批量上传总大小超过限制")
    if not candidates:
        raise ApiError(422, "CG-2603", "没有发现可导入的 CSV、XLSX、PDF、Word 或图片")

    batch_id = f"batch-{uuid.uuid4().hex}"
    jobs: list[dict[str, Any]] = []
    for index, (name, data) in enumerate(candidates, 1):
        job_id = f"import-{uuid.uuid4().hex}"
        directory = Path(".workspace") / "imports" / ctx.tenant_id / job_id
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / Path(name).name
        path.write_bytes(data)
        mode = "structured" if path.suffix.lower() in STRUCTURED_SUFFIXES else "ocr"
        recognition = _recognize_path(path)
        job = ImportJob(
            id=job_id, tenant_id=ctx.tenant_id, file_name=path.name, import_type="auto",
            status="uploaded", progress=0,
            options={"size": len(data), "path": str(path.resolve()), "mode": mode, "batchId": batch_id, "recognition": recognition, "operator": ctx.name},
            result={},
        )
        db.add(job)
        jobs.append({"jobId": job_id, "fileName": path.name, "mode": mode, "recognition": recognition, "sequence": index})
    add_audit(db, ctx, "批量上传并分类", "import", batch_id, batch_id, {"files": len(jobs), "bytes": total_bytes})
    db.commit()
    return {"batchId": batch_id, "files": jobs, "total": len(jobs), "requiresConfirmation": True}
