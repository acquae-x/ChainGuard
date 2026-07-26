from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
import csv
import hashlib
from typing import Annotated, Any

from fastapi import APIRouter, Depends
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session


from ..auth import AuthContext, require_permission
from ..database import get_db
from ..errors import ApiError
from ..models import ImportJob
from ..entity_import import DuplicateImportError, reserve_import_signature
from ..enterprise_import_catalog import IMPORT_TYPE_CATALOG
from ..import_classifier import recognize_import_type
from ..jobs import prepare_import_job
from ..repository import get_tenant_record, list_tenant_records, serialize
from ..schemas import PatchRequest


router = APIRouter(tags=["imports-settings"])

STRUCTURED_SUFFIXES = {".csv", ".xlsx"}
OCR_SUFFIXES = {".pdf", ".doc", ".docx", ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
ARCHIVE_SUFFIXES = {".zip"}


def _headers_for_path(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(next(csv.reader(handle), []))
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        sheet = load_workbook(path, read_only=True, data_only=True).active
        return [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True), ())]
    return []


def _recognize_path(path: Path) -> dict[str, Any]:
    try:
        return recognize_import_type(path.name, _headers_for_path(path))
    except Exception:
        return recognize_import_type(path.name)


def _public_job(job: ImportJob) -> dict[str, Any]:
    payload = serialize(job)
    options = dict(payload.get("options") or {})
    options.pop("path", None)
    raw_result = dict(payload.get("result") or {})
    streaming = raw_result.get("streaming") if isinstance(raw_result.get("streaming"), dict) else {}

    def count(*keys: str) -> int:
        for source in (raw_result, streaming, payload):
            for key in keys:
                if source.get(key) is not None:
                    try:
                        return int(source[key])
                    except (TypeError, ValueError):
                        continue
        return 0

    source_rows = count("total", "sourceRows")
    success_rows = count("success", "successRows", "imported")
    rejected_rows = count("failed", "rejectedRows")
    reports = raw_result.get("reports") or raw_result.get("tableReports") or streaming.get("reports") or streaming.get("tableReports") or []
    normalized_result = {
        **raw_result,
        "total": source_rows,
        "sourceRows": source_rows,
        "success": success_rows,
        "successRows": success_rows,
        "failed": rejected_rows,
        "rejectedRows": rejected_rows,
        "reports": reports,
        "tableReports": reports,
    }
    operator = payload.get("operator") or options.get("operator") or "-"
    updated_at = job.updated_at.isoformat() if job.updated_at is not None else payload.get("createdAt")
    payload.update({
        "options": options,
        "result": normalized_result,
        "total": source_rows,
        "sourceRows": source_rows,
        "success": success_rows,
        "successRows": success_rows,
        "failed": rejected_rows,
        "rejectedRows": rejected_rows,
        "reports": reports,
        "operator": operator,
        "updatedAt": updated_at,
    })
    return payload


def _camel(name: str) -> str:
    parts = name.split("_")
    return parts[0] + "".join(part.title() for part in parts[1:])


def camelize_report(report: Any) -> dict[str, Any]:
    """preflight 结果统一转 camelCase，与前端 estimatedRows/canProceed 契约对齐。"""
    data = asdict(report) if hasattr(report, "__dataclass_fields__") else dict(report)
    return {_camel(key): value for key, value in data.items()}


def normalize_xlsx_to_csv(source: str | Path) -> Path:
    """XLSX 先解析归一化为 CSV，预检必须基于真实行而不是二进制字节估算。

    解析失败必须抛出异常（由调用方转红灯），不允许吞异常后继续绿灯。
    """
    from openpyxl import load_workbook

    path = Path(source)
    sheet = load_workbook(path, read_only=True, data_only=True).active
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values, ())]
    if not any(headers):
        raise ValueError("XLSX 首行没有可用表头")
    target = path.with_suffix(".csv")
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in values:
            writer.writerow(["" if value is None else value for value in row[: len(headers)]])
    return target


def normalized_preview(path: str | Path, limit: int = 20) -> dict[str, Any]:
    """Return a bounded, structured preview of the server-side normalized file."""
    source = Path(path)
    rows: list[dict[str, Any]] = []
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if len(rows) < limit:
                    rows.append(dict(row))
                else:
                    break
    elif source.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
            sheet = load_workbook(source, read_only=True, data_only=True).active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values, ())]
            for values_row in values:
                if len(rows) >= limit:
                    break
                rows.append({headers[index]: value for index, value in enumerate(values_row) if index < len(headers) and headers[index]})
        except Exception:
            rows = []
    return {"table": source.stem, "previewRows": rows, "previewLimit": limit}

@router.post("/imports/{item_id}/preflight")
def preflight(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    from src.import_preflight import run_preflight
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id)
    suffix = Path(item.file_name).suffix.lower()
    mode = str(item.options.get("mode") or ("ocr" if suffix in OCR_SUFFIXES else "structured"))
    if suffix in OCR_SUFFIXES:
        from src.ingestion_agent import ingest_files
        intake = ingest_files([item.options["path"]])
        extraction = intake.extractions[0]
        if extraction.needs_manual:
            item.status, item.progress = "manual_required", 25
            reason = extraction.note or "未提取到可校验内容。"
            item.result = {
                "canProceed": False, "status": "manual_required",
                "message": reason,
                "recognition": item.options.get("recognition"), "extraction": asdict(extraction),
                "manualReview": {
                    "required": True, "confirmationLevel": "full", "firstImport": True,
                    "reasonCode": extraction.error_code,
                    "suggestions": ["重新上传清晰、端正且保留表头和列分隔符的图片", "改用 CSV/Excel 上传，或人工录入"],
                },
            }
            db.commit(); payload = serialize(item); payload["options"].pop("path", None); return payload
        rows = next(iter(intake.normalized.values()), [])
        if rows and all(str(key).startswith("col_") for key in rows[0]):
            detected_headers = [str(value).strip() for value in rows[0].values()]
            if len(rows) > 1 and all(detected_headers):
                rows = [
                    {detected_headers[index]: value for index, value in enumerate(source.values()) if index < len(detected_headers)}
                    for source in rows[1:]
                ]
        if not rows:
            item.status, item.progress = "manual_required", 25
            item.result = {
                "canProceed": False, "status": "manual_required",
                "message": "已识别文档，但没有形成可导入的数据行；请人工处理原文。",
                "recognition": item.options.get("recognition"), "extraction": asdict(extraction),
                "manualReview": {"required": True, "confirmationLevel": "full", "firstImport": True},
            }
            db.commit(); return _public_job(item)
        normalized_path = Path(item.options["path"]).parent / "normalized.csv"
        with normalized_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader(); writer.writerows(rows)
        recognition = recognize_import_type(item.file_name, rows[0].keys() if rows else [])
        schema_signature = hashlib.sha256("|".join(sorted(rows[0].keys() if rows else [])).encode("utf-8")).hexdigest()
        previous = [
            job for job in db.query(ImportJob).filter(ImportJob.tenant_id == ctx.tenant_id, ImportJob.status == "succeeded").all()
            if job.options.get("mode") == "ocr" and job.options.get("schemaSignature") == schema_signature
        ]
        seen_count = len(previous)
        item.options = {
            **item.options, "originalPath": item.options["path"], "path": str(normalized_path.resolve()),
            "extraction": asdict(extraction), "recognition": recognition, "schemaSignature": schema_signature,
        }
    elif suffix == ".xlsx":
        # XLSX 先归一化为 CSV 再做基于行的容量预检；解析失败必须红灯，禁止吞异常假绿灯
        try:
            normalized_path = normalize_xlsx_to_csv(item.options["path"])
        except Exception:
            item.status, item.progress = "failed", 25
            item.result = {
                "canProceed": False,
                "verdict": "PARSE_ERROR",
                "estimatedRows": None,
                "messages": ["XLSX 解析失败：文件可能损坏或不是有效的 Excel 工作簿，导入已阻止。"],
                "normalized": {"table": Path(item.file_name).stem, "previewRows": [], "previewLimit": 20},
            }
            db.commit(); payload = serialize(item); payload["options"].pop("path", None); return payload
        item.options = {**item.options, "originalPath": item.options["path"], "path": str(normalized_path.resolve())}
    report = run_preflight([item.options["path"]], Path(item.options["path"]).parent / "import.db")
    item.status = "manual_review" if mode == "ocr" and report.can_proceed else "preflighted" if report.can_proceed else "failed"
    item.progress = 25
    # camelCase 与前端 estimatedRows/canProceed 契约对齐；verdict 键名不变，confirm 闸门不受影响
    item.result = {
        **camelize_report(report), "normalized": normalized_preview(item.options["path"]),
        "recognition": item.options.get("recognition"),
    }
    if mode == "ocr":
        # Successful OCR must expose auditable engine/confidence/timing metadata
        # to the review UI; the server path remains private in job options.
        item.result["extraction"] = item.options.get("extraction")
        item.result["manualReview"] = {
            "required": True, "firstImport": seen_count == 0, "seenCount": seen_count,
            "familiarity": "novel" if seen_count == 0 else "familiar",
            "confirmationLevel": "full" if seen_count == 0 else "light",
            "confirmationPoints": ["识别出的资料类型", "关键业务字段", "低置信度/空值行"] if seen_count == 0 else ["本次差异与低置信度行"],
        }
    db.commit()
    payload = serialize(item); payload["options"].pop("path", None); return payload


@router.post("/imports/{item_id}/confirm")
def confirm_import(item_id: str, body: PatchRequest, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id)
    force = bool(body.values.get("force"))
    confirmed_type = str(body.values.get("confirmedType") or item.import_type)
    if item.import_type == "auto" and confirmed_type not in IMPORT_TYPE_CATALOG:
        raise ApiError(409, "CG-2607", "必须人工确认识别出的资料类型后才能继续")
    if confirmed_type not in IMPORT_TYPE_CATALOG:
        raise ApiError(422, "CG-2603", "资料类型不存在")
    mode = str(item.options.get("mode") or "structured")
    if mode == "ocr" and not bool(body.values.get("manualConfirmed")):
        raise ApiError(409, "CG-2608", "OCR/文档导入必须完成人工确认，不能强制绕过")
    if item.status == "failed":
        # Disk shortage is a hard safety gate, not an override-able quality warning.
        if item.result.get("verdict") == "INSUFFICIENT_DISK":
            raise ApiError(409, "CG-2602", "磁盘空间不足，不能强制导入")
        # 解析失败没有可导入的归一化数据，同样不允许强制
        if item.result.get("verdict") == "PARSE_ERROR":
            raise ApiError(409, "CG-2602", "文件解析失败，不能强制导入")
        if not force:
            raise ApiError(409, "CG-2602", "预检未通过，需明确确认后才能继续导入")
    if item.status not in {"preflighted", "manual_review", "failed"}:
        raise ApiError(409, "CG-2601", "导入任务尚未完成预检")
    item.import_type = confirmed_type
    item.status = "confirmed"; item.options = {**item.options, **body.values, "typeConfirmed": True}; item.progress = 50; db.commit()
    payload = serialize(item); payload["options"].pop("path", None); return payload


@router.post("/imports/{item_id}/execute", status_code=202)
def execute_import(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id)
    if item.status != "confirmed":
        raise ApiError(409, "CG-2601", "导入任务尚未确认")
    try:
        history = reserve_import_signature(
            db,
            ctx.tenant_id,
            item.id,
            item.file_name,
            item.options["path"],
            item.import_type,
        )
        db.flush()  # UNIQUE(tenant_id, signature) closes concurrent execute races.
    except DuplicateImportError as error:
        raise ApiError(409, "CG-2605", f"D04：相同签名文件已导入（批次 {error.history.import_job_id}）") from error
    except IntegrityError as error:
        db.rollback()
        raise ApiError(409, "CG-2605", "D04：相同签名文件已被其他导入任务占用") from error
    item.status = "pending"
    item.options = {**item.options, "signature": history.signature}
    prepare_import_job(db, item, ctx)
    db.commit()
    from . import imports_settings

    imports_settings.enqueue_import_job(item.id, ctx)
    return {"jobId": item.id, "status": item.status}


@router.get("/imports/{item_id}")
def import_progress(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id); return _public_job(item)
@router.get("/imports")
def import_history(ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]):
    data = [_public_job(item) for item in list_tenant_records(db, ImportJob, ctx.tenant_id)]
    return {"data": data, "total": len(data), "success": True}
@router.post("/imports/{item_id}/rollback")
def rollback_import(item_id: str, ctx: Annotated[AuthContext, Depends(require_permission("data:import"))], db: Annotated[Session, Depends(get_db)]): item = get_tenant_record(db, ImportJob, item_id, ctx.tenant_id); item.status = "rolled_back"; db.commit(); return {"ok": True, "id": item_id}
